"""합성 픽스처 생성(seed 고정). 담당: qa.
- sample.ifc : 2층 건물. 기둥 12(6/층), 보 16(8/층), 슬래브 2, 벽 8(4/층), 덕트 4(2/층)
- sample.dxf : 1F 평면. 레이어 GRID / A-COL / S-BEAM / A-WALL / M-DUCT. IFC와 같은 그리드, 단위 mm, 원점 오프셋+회전 적용
- sample.ply : 1F 스캔 합성. 기둥 6개 중 3개 완전, 1개 절반, 1개 offset 80mm, 1개 가림(점 없음)
- alignment.json : 기준점 3점(스캔↔모델)
- schedule.csv / schedule.xml / schedule.xer : 1F 기둥→보→슬래브 작업
- *.expected.json : 기대값

재현성(바이트 동일): 기하·점군·판정은 numpy seed 42. 그 외 라이브러리가 매 실행 바꾸는 값도 고정한다 —
IFC GlobalId(uuid4 → seed 고정 생성기), IFC FILE_NAME 타임스탬프, IfcRel* 집합 순서(express id 정렬),
DXF $TDCREATE/$TDUPDATE/$VERSIONGUID/$FINGERPRINTGUID(ezdxf 고정 메타데이터 옵션).
"""
from __future__ import annotations

import json
import os
import re
import sys
import uuid
import zipfile
from datetime import datetime
from pathlib import Path

import ezdxf
import ifcopenshell
import ifcopenshell.api
import ifcopenshell.guid
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Border, Side

OUT = Path(__file__).resolve().parents[1]
rng = np.random.default_rng(42)

# GlobalId 를 결정적으로: ifcopenshell.api root.create_entity 는 ifcopenshell.guid.new()(uuid4)를 호출한다.
GUID_SEED = 4242
FIXED_TIMESTAMP = "2026-01-01T00:00:00+00:00"
_guid_rng = np.random.default_rng(GUID_SEED)


def _seeded_guid() -> str:
    return ifcopenshell.guid.compress(uuid.UUID(bytes=_guid_rng.bytes(16), version=4).hex)


ifcopenshell.guid.new = _seeded_guid
ezdxf.options.write_fixed_meta_data_for_testing = True   # $TDCREATE/$TDUPDATE/$VERSIONGUID/$FINGERPRINTGUID 고정

GRID_X = [0.0, 6.0, 12.0]        # m
GRID_Y = [0.0, 8.0]
LEVELS = [("1F", 0.0), ("2F", 4.0)]
COL = 0.6                        # 기둥 단면 m
SLAB_T = 0.2
STOREY_H = 4.0

# DXF 좌표계: 원점 오프셋 (m) + 회전(도). sync-2d3d의 정합 테스트용
DXF_ORIGIN = (100.0, 50.0)
DXF_ROT_DEG = 15.0
DXF_UNIT_TO_M = 0.001            # mm

# 스캔 좌표계: 모델 좌표에 rigid transform 적용
SCAN_ROT_DEG = -7.0
SCAN_T = np.array([3.2, -1.5, 0.3])


def _rot2(deg: float) -> np.ndarray:
    t = np.radians(deg)
    return np.array([[np.cos(t), -np.sin(t)], [np.sin(t), np.cos(t)]])


def model_to_dxf(xy: np.ndarray) -> np.ndarray:
    """모델(m) → DXF(mm)"""
    return ((_rot2(DXF_ROT_DEG) @ xy.T).T + np.array(DXF_ORIGIN)) / DXF_UNIT_TO_M


def model_to_scan(xyz: np.ndarray) -> np.ndarray:
    r = np.eye(3)
    r[:2, :2] = _rot2(SCAN_ROT_DEG)
    return (r @ xyz.T).T + SCAN_T


# ---------------------------------------------------------------- IFC
def build_ifc() -> dict:
    f = ifcopenshell.api.run("project.create_file", version="IFC4")
    project = ifcopenshell.api.run("root.create_entity", f, ifc_class="IfcProject", name="BuildTwin Sample")
    ifcopenshell.api.run("unit.assign_unit", f, length={"is_metric": True, "raw": "METERS"})
    ctx = ifcopenshell.api.run("context.add_context", f, context_type="Model")
    body = ifcopenshell.api.run("context.add_context", f, context_type="Model", context_identifier="Body",
                                target_view="MODEL_VIEW", parent=ctx)
    site = ifcopenshell.api.run("root.create_entity", f, ifc_class="IfcSite", name="Site")
    bldg = ifcopenshell.api.run("root.create_entity", f, ifc_class="IfcBuilding", name="Building")
    ifcopenshell.api.run("aggregate.assign_object", f, relating_object=project, products=[site])
    ifcopenshell.api.run("aggregate.assign_object", f, relating_object=site, products=[bldg])

    expected = {"IfcColumn": 0, "IfcBeam": 0, "IfcSlab": 0, "IfcWall": 0, "IfcDuctSegment": 0}
    ids: dict[str, list[dict]] = {"columns": [], "beams": [], "slabs": [], "walls": [], "ducts": []}

    def add(ifc_class: str, name: str, storey, origin, size, key: str, extra: dict | None = None, rot_deg: float = 0.0):
        """origin: 배치 원점(월드, m). size: (length, thickness, height) 로컬 축 기준. rot_deg: Z축 회전.
        회전 시 로컬 +x → 월드 방향이 회전하므로 월드 bbox를 함께 계산해 expected에 기록한다."""
        el = ifcopenshell.api.run("root.create_entity", f, ifc_class=ifc_class, name=name)
        ifcopenshell.api.run("spatial.assign_container", f, relating_structure=storey, products=[el])
        mat = np.eye(4)
        mat[:2, :2] = _rot2(rot_deg)
        mat[:3, 3] = origin
        ifcopenshell.api.run("geometry.edit_object_placement", f, product=el, matrix=mat)
        rep = ifcopenshell.api.run("geometry.add_wall_representation", f, context=body,
                                   length=size[0], height=size[2], thickness=size[1])
        ifcopenshell.api.run("geometry.assign_representation", f, product=el, representation=rep)
        pset = ifcopenshell.api.run("pset.add_pset", f, product=el, name="Pset_BuildTwin")
        ifcopenshell.api.run("pset.edit_pset", f, pset=pset, properties={"Zone": extra.get("zone", "Z1") if extra else "Z1"})
        # 월드 bbox
        local = np.array([[0, 0, 0], [size[0], 0, 0], [size[0], size[1], 0], [0, size[1], 0]], dtype=float)
        world = (mat[:3, :3] @ local.T).T + np.array(origin)
        bmin = [float(world[:, 0].min()), float(world[:, 1].min()), float(origin[2])]
        bmax = [float(world[:, 0].max()), float(world[:, 1].max()), float(origin[2] + size[2])]
        expected[ifc_class] += 1
        ids[key].append({"global_id": el.GlobalId, "name": name, "origin": list(map(float, origin)),
                         "size": list(map(float, size)), "rot_deg": rot_deg, "bbox": [bmin, bmax], **(extra or {})})
        return el

    for li, (lname, elev) in enumerate(LEVELS):
        storey = ifcopenshell.api.run("root.create_entity", f, ifc_class="IfcBuildingStorey", name=lname)
        storey.Elevation = elev
        ifcopenshell.api.run("aggregate.assign_object", f, relating_object=bldg, products=[storey])
        ifcopenshell.api.run("geometry.edit_object_placement", f, product=storey,
                             matrix=np.array([[1, 0, 0, 0], [0, 1, 0, 0], [0, 0, 1, elev], [0, 0, 0, 1]], dtype=float))
        z0 = elev
        # 기둥 6개 (3x2 그리드)
        for xi, gx in enumerate(GRID_X):
            for yi, gy in enumerate(GRID_Y):
                add("IfcColumn", f"C{li+1}-{xi+1}{yi+1}", storey, (gx - COL / 2, gy - COL / 2, z0),
                    (COL, COL, STOREY_H - SLAB_T), "columns", {"level": lname, "grid": f"{xi}{yi}"})
        # 보 8개: X방향 4 + Y방향 3 + 캔틸레버 1
        z = z0 + STOREY_H - SLAB_T - 0.5
        for yi, gy in enumerate(GRID_Y):
            for xi in range(len(GRID_X) - 1):
                x0 = GRID_X[xi] + COL / 2
                add("IfcBeam", f"BX{li+1}-{xi+1}{yi+1}", storey, (x0, gy - 0.15, z),
                    (GRID_X[xi + 1] - COL / 2 - x0, 0.3, 0.5), "beams", {"level": lname})
        for xi, gx in enumerate(GRID_X):
            y0 = GRID_Y[0] + COL / 2
            # 로컬 +x → 월드 +y (90° 회전). 로컬 +y(두께) → 월드 -x. 원점을 gx+0.15에 두면 x∈[gx-0.15, gx+0.15]
            add("IfcBeam", f"BY{li+1}-{xi+1}", storey, (gx + 0.15, y0, z), (GRID_Y[1] - COL, 0.3, 0.5), "beams",
                {"level": lname}, rot_deg=90)
        add("IfcBeam", f"BC{li+1}", storey, (GRID_X[-1] + COL / 2, GRID_Y[0] - 0.15, z), (1.5, 0.3, 0.5), "beams",
            {"level": lname, "cantilever": True})
        # 슬래브 1
        add("IfcSlab", f"S{li+1}", storey, (-0.5, -0.5, z0 + STOREY_H - SLAB_T), (GRID_X[-1] + 1.0, GRID_Y[-1] + 1.0, SLAB_T), "slabs", {"level": lname})
        # 벽 4 (외벽)
        add("IfcWall", f"W{li+1}-S", storey, (0.0, -0.4, z0), (GRID_X[-1], 0.2, STOREY_H - SLAB_T), "walls", {"level": lname})
        add("IfcWall", f"W{li+1}-N", storey, (0.0, GRID_Y[-1] + 0.2, z0), (GRID_X[-1], 0.2, STOREY_H - SLAB_T), "walls", {"level": lname})
        for tag, gx in (("W", GRID_X[0] - 0.4), ("E", GRID_X[-1] + 0.2)):
            # 90° 회전: 원점 x=gx+0.2 → 월드 x∈[gx, gx+0.2]
            add("IfcWall", f"W{li+1}-{tag}", storey, (gx + 0.2, 0.0, z0), (GRID_Y[-1], 0.2, STOREY_H - SLAB_T), "walls",
                {"level": lname}, rot_deg=90)
        # 덕트 2
        for di, gy in enumerate((2.0, 6.0)):
            add("IfcDuctSegment", f"D{li+1}-{di+1}", storey, (0.5, gy - 0.2, z0 + STOREY_H - SLAB_T - 1.2), (GRID_X[-1] - 1.0, 0.4, 0.4), "ducts", {"level": lname})

    # 집합(set) 순서로 쓰이는 관계 속성은 express id 로 정렬해 파일 바이트를 고정한다(의미 불변)
    for rel in f.by_type("IfcRelContainedInSpatialStructure"):
        rel.RelatedElements = sorted(rel.RelatedElements, key=lambda e: e.id())
    for rel in f.by_type("IfcRelAggregates"):
        rel.RelatedObjects = sorted(rel.RelatedObjects, key=lambda e: e.id())
    for ua in f.by_type("IfcUnitAssignment"):          # unit.assign_unit 이 set 으로 만들어 id() 순서가 프로세스마다 다르다
        ua.Units = sorted(ua.Units, key=lambda e: e.id())
    for rel in f.by_type("IfcRelDefinesByProperties"):
        rel.RelatedObjects = sorted(rel.RelatedObjects, key=lambda e: e.id())
    f.header.file_name.time_stamp = FIXED_TIMESTAMP
    f.write(str(OUT / "sample.ifc"))
    (OUT / "sample.ifc.expected.json").write_text(json.dumps({
        "counts": expected, "levels": [{"name": n, "elevation": e} for n, e in LEVELS],
        "objects": ids, "grid_x": GRID_X, "grid_y": GRID_Y, "unit": "m",
    }, indent=2, ensure_ascii=False))
    return ids


# ---------------------------------------------------------------- DXF
def build_dxf(ids: dict) -> None:
    doc = ezdxf.new("R2010")
    doc.header["$INSUNITS"] = 4   # mm
    msp = doc.modelspace()
    for name, color in (("GRID", 8), ("A-COL", 1), ("S-BEAM", 3), ("A-WALL", 5), ("M-DUCT", 6), ("A-TEXT", 7)):
        doc.layers.add(name, color=color)
    expected_entities: dict[str, int] = {}
    mapping_expected: list[dict] = []

    def line(p0, p1, layer):
        e = msp.add_line(tuple(model_to_dxf(np.array([p0]))[0]), tuple(model_to_dxf(np.array([p1]))[0]), dxfattribs={"layer": layer})
        expected_entities[layer] = expected_entities.get(layer, 0) + 1
        return e

    def rect(x0, y0, w, h, layer, closed=True):
        pts = model_to_dxf(np.array([[x0, y0], [x0 + w, y0], [x0 + w, y0 + h], [x0, y0 + h]]))
        e = msp.add_lwpolyline([tuple(p) for p in pts], close=closed, dxfattribs={"layer": layer})
        expected_entities[layer] = expected_entities.get(layer, 0) + 1
        return e

    # 그리드선
    for gx in GRID_X:
        line((gx, -1.5), (gx, GRID_Y[-1] + 1.5), "GRID")
    for gy in GRID_Y:
        line((-1.5, gy), (GRID_X[-1] + 1.5, gy), "GRID")
    # 1F 기둥 → A-COL 폴리라인 (매핑 기대값)
    for c in ids["columns"]:
        if c["level"] != "1F":
            continue
        e = rect(c["origin"][0], c["origin"][1], c["size"][0], c["size"][1], "A-COL")
        mapping_expected.append({"handle": e.dxf.handle, "global_id": c["global_id"], "layer": "A-COL"})
    # 1F 보 → S-BEAM 라인(중심선)
    for b in ids["beams"]:
        if b["level"] != "1F":
            continue
        (x0, y0, _), (x1, y1, _) = b["bbox"]
        if b["name"].startswith("BY"):
            e = line(((x0 + x1) / 2, y0), ((x0 + x1) / 2, y1), "S-BEAM")
        else:
            e = line((x0, (y0 + y1) / 2), (x1, (y0 + y1) / 2), "S-BEAM")
        mapping_expected.append({"handle": e.dxf.handle, "global_id": b["global_id"], "layer": "S-BEAM"})
    # 1F 벽
    for w in ids["walls"]:
        if w["level"] != "1F":
            continue
        (x0, y0, _), (x1, y1, _) = w["bbox"]
        e = rect(x0, y0, x1 - x0, y1 - y0, "A-WALL")
        mapping_expected.append({"handle": e.dxf.handle, "global_id": w["global_id"], "layer": "A-WALL"})
    # 1F 덕트
    for d in ids["ducts"]:
        if d["level"] != "1F":
            continue
        o, s = d["origin"], d["size"]
        e = rect(o[0], o[1], s[0], s[1], "M-DUCT")
        mapping_expected.append({"handle": e.dxf.handle, "global_id": d["global_id"], "layer": "M-DUCT"})
    # 타이틀 텍스트
    t = msp.add_text("1F PLAN", dxfattribs={"layer": "A-TEXT", "height": 300})
    t.set_placement(tuple(model_to_dxf(np.array([[0.0, -1.2]]))[0]))
    expected_entities["A-TEXT"] = 1
    # 블록 참조 1개 (기둥 심볼 블록) — 2F용 참고
    blk = doc.blocks.new("COL_SYM")
    blk.add_circle((0, 0), 200)
    ins = msp.add_blockref("COL_SYM", tuple(model_to_dxf(np.array([[GRID_X[1], GRID_Y[0]]]))[0]), dxfattribs={"layer": "A-COL"})
    expected_entities["A-COL"] += 1
    doc.saveas(str(OUT / "sample.dxf"))
    (OUT / "sample.dxf.expected.json").write_text(json.dumps({
        "entity_counts_by_layer": expected_entities, "insunits": 4, "unit_to_m": DXF_UNIT_TO_M,
        "alignment": {"origin_m": DXF_ORIGIN, "rotation_deg": DXF_ROT_DEG, "scale": DXF_UNIT_TO_M},
        "block_insert_handle": ins.dxf.handle,
    }, indent=2))
    (OUT / "mapping.expected.json").write_text(json.dumps({"level": "1F", "mappings": mapping_expected}, indent=2))


# ---------------------------------------------------------------- PLY (scan)
def _sample_box_surface(origin, size, density_per_m2: float) -> np.ndarray:
    o, s = np.array(origin), np.array(size)
    pts = []
    faces = [  # (axis fixed, value, other two axes)
        (0, o[0]), (0, o[0] + s[0]), (1, o[1]), (1, o[1] + s[1]), (2, o[2]), (2, o[2] + s[2]),
    ]
    for ax, val in faces:
        others = [i for i in range(3) if i != ax]
        area = s[others[0]] * s[others[1]]
        n = max(1, int(area * density_per_m2))
        p = np.zeros((n, 3))
        p[:, ax] = val
        for i in others:
            p[:, i] = rng.uniform(o[i], o[i] + s[i], n)
        pts.append(p)
    return np.vstack(pts)


def build_ply(ids: dict) -> None:
    pts = []
    verdict_expected: dict[str, str] = {}
    cols = [c for c in ids["columns"] if c["level"] == "1F"]
    # 0,1,2 완료 / 3 절반 / 4 offset / 5 가림(점 없음)
    for i, c in enumerate(cols):
        o, s = c["origin"], c["size"]
        if i <= 2:
            pts.append(_sample_box_surface(o, s, 800))
            verdict_expected[c["global_id"]] = "ESTIMATED_DONE"
        elif i == 3:
            pts.append(_sample_box_surface(o, (s[0], s[1], s[2] * 0.45), 800))
            verdict_expected[c["global_id"]] = "IN_PROGRESS"
        elif i == 4:
            pts.append(_sample_box_surface((o[0] + 0.08, o[1] + 0.08, o[2]), s, 800))
            verdict_expected[c["global_id"]] = "MISMATCH"
        else:
            verdict_expected[c["global_id"]] = "UNVERIFIABLE"   # 가림: 스캐너 앞에 차폐 박스
    # 차폐 박스: 마지막 기둥과 스캐너 사이
    occl = cols[5]
    ox, oy = occl["origin"][0], occl["origin"][1]
    pts.append(_sample_box_surface((ox - 1.5, oy - 0.6, 0.0), (0.3, 1.8, 3.0), 600))
    # 1F 바닥 슬래브 상면(참조용 대면적)
    pts.append(_sample_box_surface((-0.5, -0.5, -0.02), (GRID_X[-1] + 1.0, GRID_Y[-1] + 1.0, 0.02), 150))
    # 벽 1F 남측 완료
    w = [w for w in ids["walls"] if w["level"] == "1F" and w["name"].endswith("-S")][0]
    (wx0, wy0, wz0), (wx1, wy1, wz1) = w["bbox"]
    pts.append(_sample_box_surface((wx0, wy0, wz0), (wx1 - wx0, wy1 - wy0, wz1 - wz0), 500))
    verdict_expected[w["global_id"]] = "ESTIMATED_DONE"
    for w in ids["walls"]:
        if w["level"] == "1F" and w["global_id"] not in verdict_expected:
            verdict_expected[w["global_id"]] = "NOT_BUILT"
    for b in ids["beams"] + ids["slabs"] + ids["ducts"]:
        if b["level"] == "1F":
            verdict_expected[b["global_id"]] = "NOT_BUILT"
    model_pts = np.vstack(pts)
    model_pts += rng.normal(0, 0.004, model_pts.shape)   # 4mm 노이즈
    scan_pts = model_to_scan(model_pts)
    # PLY ascii
    with open(OUT / "sample.ply", "w") as fh:
        fh.write(f"ply\nformat ascii 1.0\nelement vertex {len(scan_pts)}\nproperty float x\nproperty float y\nproperty float z\nend_header\n")
        np.savetxt(fh, scan_pts, fmt="%.4f")
    scanner_model = np.array([[ox - 4.0, oy + 0.3, 1.5]])
    cps_model = np.array([[0.0, 0.0, 0.0], [12.0, 0.0, 0.0], [12.0, 8.0, 0.0], [0.0, 8.0, 3.8]])
    cps_scan = model_to_scan(cps_model)
    (OUT / "alignment.json").write_text(json.dumps({
        "control_points": [{"name": f"CP{i+1}", "scan_xyz": cps_scan[i].round(4).tolist(), "model_xyz": cps_model[i].tolist()} for i in range(4)],
        "scanner_position": model_to_scan(scanner_model)[0].round(4).tolist(),
    }, indent=2))
    (OUT / "verdict.expected.json").write_text(json.dumps({
        "scan_rotation_deg": SCAN_ROT_DEG, "scan_translation": SCAN_T.tolist(), "level": "1F",
        "verdicts": verdict_expected, "point_count": int(len(scan_pts)),
    }, indent=2))


# ---------------------------------------------------------------- Schedule
def build_schedules(ids: dict) -> None:
    acts = [
        ("A100", "1F 기둥 철근·거푸집·타설", "1.1.1", "structure", "1F", "Z1", "2026-09-01", "2026-09-10", 8),
        ("A110", "1F 보 시공", "1.1.2", "structure", "1F", "Z1", "2026-09-11", "2026-09-18", 6),
        ("A120", "1F 슬래브 타설", "1.1.3", "structure", "1F", "Z1", "2026-09-19", "2026-09-24", 4),
        ("A200", "1F 외벽 조적", "1.2.1", "architecture", "1F", "Z1", "2026-09-25", "2026-10-05", 8),
        ("A300", "1F 덕트 설치", "1.3.1", "mechanical", "1F", "Z1", "2026-10-06", "2026-10-12", 5),
        ("A400", "2F 기둥 시공", "2.1.1", "structure", "2F", "Z1", "2026-09-25", "2026-10-04", 8),
    ]
    rels = [("A100", "A110", "FS", 0), ("A110", "A120", "FS", 0), ("A120", "A200", "FS", 0), ("A120", "A300", "FS", 2), ("A120", "A400", "FS", 0)]
    with open(OUT / "schedule.csv", "w", encoding="utf-8") as fh:
        fh.write("activity_id,name,wbs_code,discipline,level,zone,planned_start,planned_finish,duration_days,predecessors,crew\n")
        for a in acts:
            preds = ";".join(f"{p}:{t}:{lag}" for p, s, t, lag in rels if s == a[0])
            fh.write(",".join(map(str, a)) + f",{preds},4\n")
    # MS Project XML (최소)
    xml = ['<?xml version="1.0" encoding="UTF-8"?>', '<Project xmlns="http://schemas.microsoft.com/project"><Tasks>']
    for i, a in enumerate(acts, 1):
        xml.append(f"<Task><UID>{i}</UID><ID>{i}</ID><Name>{a[1]}</Name><WBS>{a[2]}</WBS><Start>{a[6]}T08:00:00</Start><Finish>{a[7]}T17:00:00</Finish><Duration>PT{a[8]*8}H0M0S</Duration>")
        xml.append("<PredecessorLink>" + "</PredecessorLink><PredecessorLink>".join(
            f"<PredecessorUID>{[x[0] for x in acts].index(p)+1}</PredecessorUID><Type>{ {'FS':1,'SS':3,'FF':0,'SF':2}[t] }</Type><LinkLag>{lag*4800}</LinkLag>"
            for p, s, t, lag in rels if s == a[0]) + "</PredecessorLink>" if any(s == a[0] for _, s, _, _ in rels) else "")
        xml.append(f"<ExtendedAttribute><FieldID>188743731</FieldID><Value>{a[0]}</Value></ExtendedAttribute></Task>")
    xml.append("</Tasks></Project>")
    (OUT / "schedule.xml").write_text("\n".join(xml).replace("<PredecessorLink></PredecessorLink>", ""))
    # P6 XER (최소 테이블)
    lines = ["ERMHDR\t20.12\t2026-09-01\tProject\tadmin\tadmin\tdbxDatabaseNoName\tProject Management\tKRW",
             "%T\tPROJECT", "%F\tproj_id\tproj_short_name", "%R\t1\tBT",
             "%T\tTASK", "%F\ttask_id\tproj_id\ttask_code\ttask_name\twbs_id\ttarget_start_date\ttarget_end_date\ttarget_drtn_hr_cnt\tphys_complete_pct"]
    for i, a in enumerate(acts, 1):
        lines.append(f"%R\t{i}\t1\t{a[0]}\t{a[1]}\t{a[2]}\t{a[6]} 08:00\t{a[7]} 17:00\t{a[8]*8}\t0")
    lines += ["%T\tTASKPRED", "%F\ttask_pred_id\ttask_id\tpred_task_id\tpred_type\tlag_hr_cnt"]
    codes = [x[0] for x in acts]
    for j, (p, s, t, lag) in enumerate(rels, 1):
        lines.append(f"%R\t{j}\t{codes.index(s)+1}\t{codes.index(p)+1}\tPR_{t}\t{lag*8}")
    lines.append("%E")
    (OUT / "schedule.xer").write_text("\n".join(lines), encoding="utf-8")
    (OUT / "schedule.expected.json").write_text(json.dumps({
        "activity_count": len(acts), "relation_count": len(rels),
        "activities": [a[0] for a in acts], "relations": [[p, s, t, lag] for p, s, t, lag in rels],
        "activity_object_mapping": {"A100": [c["global_id"] for c in ids["columns"] if c["level"] == "1F"],
                                    "A110": [b["global_id"] for b in ids["beams"] if b["level"] == "1F"],
                                    "A120": [s["global_id"] for s in ids["slabs"] if s["level"] == "1F"],
                                    "A200": [w["global_id"] for w in ids["walls"] if w["level"] == "1F"],
                                    "A300": [d["global_id"] for d in ids["ducts"] if d["level"] == "1F"],
                                    "A400": [c["global_id"] for c in ids["columns"] if c["level"] == "2F"]},
    }, indent=2))


# --------------------------------------------------------------------------- 문서관리대장 (ADR 0007)
# 실제 현장(고창CDC) 대장 형식을 그대로 재현한다. 깨끗한 대장으로 테스트하면 실물에서 무너지므로,
# 현장에서 실제로 관찰된 지저분함을 의도적으로 심는다 — 아래 MESS-1..7 주석 참고.
REGISTER_TITLE = "(고창 CDC 물류센터 신축공사)"

# 시트마다 컬럼 위치가 다르다(실제 대장에서 TFA는 제목이 H열, TFR은 G열). 열 위치를 상수로 박으면 안 되고
# 3행 헤더를 읽어 찾아야 한다는 것을 이 픽스처가 강제한다.
REGISTER_SHEETS = {
    "TFA": {"title": "승인/검토/참조 요청서(TFA)",
            "headers": ["No", "문서발생일", "발신", "공종", "번호", "문서번호", "회신요청일", "제목", "처리결과", "처리완료일"]},
    "TFR": {"title": "자료제출서(TFR)",
            "headers": ["No", "문서발생일", "발신", "공종", "번호", "문서번호", "제목", "처리결과", "처리완료일"]},
}

# (발생일, 발신, 공종열, 번호, 문서번호원문, 제목, 처리결과, 처리완료일)
TFA_ROWS = [
    ("26-09-01", "동부", "구조", "26049", "동부-HG-TFA-구조-26-049",
     "시공상세도 승인요청 - 1F 기둥 배근도 (Z1)", "승인", "26-09-03"),
    ("26-09-02", "동부", "구조", "26050", "동부-HG-TFA-구조-26-050",
     "시공상세도 승인요청 - 1F 보 배근도 (Z1)", "조건부 승인", "26-09-05"),
    # MESS-1: 처리결과가 공란. 실무에서 흔하다. UNKNOWN 이어야 하며 절대 승인으로 추측하면 안 된다.
    ("26-09-04", "동부", "구조", "26051", "동부-HG-TFA-구조-26-051",
     "시공상세도 승인요청 - 1F 슬래브 배근도 (Z1)", None, None),
    # MESS-2: 문서번호의 공종 토큰(토목)과 공종 열(건축)이 어긋난다. 협력사 입력 편차 — 실제로 자주 발생.
    #         문서번호의 공종을 신뢰해 매핑하면 여기서 틀린다(ADR 0007).
    ("26-09-08", "동부", "건축", "26011", "동부-HG-TFA-토목-26-011",
     "자재승인원 - 외벽 조적 벽돌 (1F Z1)", "반려", "26-09-11"),
    # MESS-3: 처리결과에 앞뒤 공백. strip 안 하면 정규화 표에서 미스.
    ("26-09-10", "중원", "기계", "26023", "중원-HG-TFA-기계-26023",
     "시공상세도 승인요청 - 1F 덕트 경로도 (Z1)", "  검토중  ", None),
    ("26-09-12", "동부", "구조", "26052", "동부-HG-TFA-구조-26-052",
     "시공상세도 승인요청 - 2F 기둥 배근도 (Z1)", "승인", "26-09-15"),
    # MESS-4: 제목이 위 1F 기둥 건과 ZONE 하나만 다르다. 제목 유사도는 0.9를 넘지만 별개 문서다.
    #         "유사도 높으면 자동 확정" 을 하면 안 된다는 규칙(ADR 0007)을 이 행이 검증한다.
    ("26-09-16", "동부", "구조", "26053", "동부-HG-TFA-구조-26-053",
     "시공상세도 승인요청 - 1F 기둥 배근도 (Z2)", "승인", "26-09-18"),
    # MESS-5: 공종이 두 토큰(통신-품질), 번호가 날짜형(제26-07-09호 → 260709). 자릿수를 재해석하면 안 된다.
    ("26-09-18", "S1", "통신", "260709", "S1-HG-TFA-통신-품질-제26-07-09호",
     "통신 배관 경로 검토요청 - 1F", "재제출 요청", "26-09-20"),
]

TFR_ROWS = [
    ("26-09-05", "동부", "구조", "26007", "동부-HG-TFR-구조-26-007",
     "1F 기둥 콘크리트 배합설계서 제출", "접수", "26-09-06"),
    ("26-09-09", "중원", "기계", "26004", "중원-HG-TFR-기계-26004",
     "덕트 자재 시험성적서 제출", None, None),
]


def _register_sheet(wb, key: str, rows: list) -> None:
    spec = REGISTER_SHEETS[key]
    ws = wb.create_sheet(key)
    ncol = len(spec["headers"])
    last = chr(ord("A") + ncol - 1)
    # 1~2행: 병합된 제목. 실제 대장과 같게 두어 헤더 탐색이 3행을 찾아내는지 검증한다.
    ws.merge_cells(f"A1:{last}1")
    ws["A1"] = spec["title"]
    ws.merge_cells(f"A2:{last}2")
    ws["A2"] = REGISTER_TITLE
    for col, name in enumerate(spec["headers"], start=1):
        ws.cell(row=3, column=col, value=name)

    idx = {name: i + 1 for i, name in enumerate(spec["headers"])}
    for n, (issued, sender, discipline, number, docno, title, result, decided) in enumerate(rows, start=1):
        r = 3 + n
        ws.cell(row=r, column=idx["No"], value=n)
        # MESS-6: 발생일이 어떤 행은 문자열, 어떤 행은 날짜값. 실제 대장에서 섞여 있다.
        ws.cell(row=r, column=idx["문서발생일"],
                value=datetime.strptime(issued, "%y-%m-%d") if n % 3 == 0 else issued)
        ws.cell(row=r, column=idx["발신"], value=sender)
        ws.cell(row=r, column=idx["공종"], value=discipline)
        ws.cell(row=r, column=idx["번호"], value=number)
        # 문서번호(F열)는 실제 대장에서 수식으로 만들어진다. 값이 아니라 수식이 들어 있는 경우를 재현한다.
        ws.cell(row=r, column=idx["문서번호"], value=docno)
        ws.cell(row=r, column=idx["제목"], value=title)
        if result is not None:
            ws.cell(row=r, column=idx["처리결과"], value=result)
        if decided is not None:
            ws.cell(row=r, column=idx["처리완료일"], value=decided)
        if "회신요청일" in idx:
            ws.cell(row=r, column=idx["회신요청일"], value=None)

    # MESS-7: 데이터 끝난 뒤 서식만 있고 값은 없는 빈 행들. 마지막 데이터 행을 행 개수로 판단하면 틀린다.
    for r in range(4 + len(rows), 4 + len(rows) + 12):
        for col in range(1, ncol + 1):
            ws.cell(row=r, column=col).border = Border(bottom=Side(style="thin"))


def _normalize_xlsx(path: Path) -> None:
    """xlsx 를 바이트 재현 가능하게 만든다.

    openpyxl 은 zip 항목마다 현재 시각을 쓰므로 매 실행 파일이 달라진다. 다른 픽스처는 전부
    바이트 동일하게 재생성되므로(모듈 상단 참고) 대장만 예외로 둘 수 없다. 항목 순서를 이름으로
    정렬하고 타임스탬프를 고정해 다시 쓴다. 내용은 건드리지 않는다.
    """
    fixed = (2026, 1, 1, 0, 0, 0)
    stamp = "2026-01-01T00:00:00Z"
    with zipfile.ZipFile(path) as src:
        items = sorted(src.infolist(), key=lambda i: i.filename)
        payload = [(i.filename, src.read(i.filename)) for i in items]
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as dst:
        for name, data in payload:
            if name == "docProps/core.xml":
                # openpyxl 은 wb.properties.modified 에 무엇을 넣든 저장 시점의 현재 시각으로 덮어쓴다.
                data = re.sub(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                              rb"\g<1>" + stamp.encode() + rb"\g<2>", data)
            info = zipfile.ZipInfo(name, date_time=fixed)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            dst.writestr(info, data)


def build_document_register() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _register_sheet(wb, "TFA", TFA_ROWS)
    _register_sheet(wb, "TFR", TFR_ROWS)
    # 비어 있는 다른 종류 시트. 시트 탐색이 데이터 없는 시트에서 넘어지지 않는지 본다.
    empty = wb.create_sheet("NCR")
    empty["A1"] = "부적합 보고서(NCR)"
    # 재현성: openpyxl 이 매번 현재 시각을 docProps 에 쓰므로 고정한다.
    fixed = datetime(2026, 1, 1, 0, 0, 0)
    wb.properties.created = fixed
    wb.properties.modified = fixed
    wb.properties.creator = "buildtwin-fixtures"
    wb.properties.lastModifiedBy = "buildtwin-fixtures"
    wb.save(OUT / "document_register.xlsx")
    _normalize_xlsx(OUT / "document_register.xlsx")

    (OUT / "document_register.expected.json").write_text(json.dumps({
        "sheets": {"TFA": len(TFA_ROWS), "TFR": len(TFR_ROWS), "NCR": 0},
        "header_row": 3,
        "first_data_row": 4,
        # 시트마다 제목·처리결과 열 위치가 다르다는 사실 자체가 기대값이다.
        "title_column": {"TFA": "H", "TFR": "G"},
        "result_column": {"TFA": "I", "TFR": "H"},
        "trailing_formatted_blank_rows": 12,
        "cases": {
            "blank_result_is_unknown": "동부-HG-TFA-구조-26-051",
            "discipline_token_conflicts_with_column": {
                "doc_no": "동부-HG-TFA-토목-26-011", "token": "토목", "column": "건축"},
            "result_needs_strip": "중원-HG-TFA-기계-26023",
            "near_duplicate_title_different_zone": [
                "동부-HG-TFA-구조-26-049", "동부-HG-TFA-구조-26-053"],
            "two_token_discipline_and_date_number": {
                "doc_no": "S1-HG-TFA-통신-품질-제26-07-09호", "discipline": "통신-품질", "number": "260709"},
            "date_cell_types_mixed": True,
        },
    }, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    # ifcopenshell/ezdxf 내부의 set 순회가 문자열 해시 랜덤화에 좌우되므로(파일 바이트 순서), 해시 시드를 고정해 재실행한다.
    if os.environ.get("PYTHONHASHSEED") != "0":
        os.execve(sys.executable, [sys.executable, __file__, *sys.argv[1:]], {**os.environ, "PYTHONHASHSEED": "0"})
    ids = build_ifc()
    build_dxf(ids)
    build_ply(ids)
    build_schedules(ids)
    build_document_register()
    print("fixtures written to", OUT)
    print(json.dumps({k: len(v) for k, v in ids.items()}))
    sys.exit(0)
