"""리뷰어 재현 — 임시 파일. 리뷰 종료 시 삭제한다."""
from __future__ import annotations

import copy
from pathlib import Path
from typing import Any

import openpyxl
import pytest
import yaml

from packages.core.settings import settings
from services.progress.config_loader import load_config

from .conftest import FIXTURES, add_member, upload

REGISTER = FIXTURES / "document_register.xlsx"
SCHEDULE = FIXTURES / "schedule.csv"
_TFA = {"no": 1, "issued": 2, "sender": 3, "discipline": 4, "seq": 5, "doc_number": 6,
        "reply_due": 7, "title": 8, "result": 9, "completed": 10}
_TFA_FIRST_FREE_ROW = 12

ACT = "A200"
TITLE = "시공상세도 승인요청 - 1F 외벽 조적 상세도 (Z1)"
SEQ = 26210
OLD = {"sender": "중원", "discipline": "건축", "seq": SEQ, "result": "반려",
       "doc_number": "중원-HG-TFA-건축-26-210", "title": TITLE}
NEW = {"sender": "중원E&C", "discipline": "건축", "seq": SEQ, "result": "승인",
       "doc_number": "중원EC-HG-TFA-건축-26-211", "title": TITLE}


def _new_project(client, auth, user_ids, name: str) -> str:
    r = client.post("/api/projects", headers=auth("admin"), json={"name": name})
    assert r.status_code == 201, r.text
    pid = r.json()["project_id"]
    for role in ("contractor", "cm", "client"):
        add_member(client, auth("admin"), pid, user_ids[role], role)
    return pid


def _write_cfg(target: Path, mutate) -> Path:
    cfg = copy.deepcopy(load_config("document_register.yaml"))
    mutate(cfg)
    target.mkdir(parents=True, exist_ok=True)
    (target / "document_register.yaml").write_text(yaml.safe_dump(cfg, allow_unicode=True), encoding="utf-8")
    return target


def _upload_cfg(client, headers, pid, path, config_dir):
    prev = settings.config_dir
    settings.config_dir = str(config_dir)
    try:
        return upload(client, headers, pid, path)
    finally:
        settings.config_dir = prev


def _register(dest: Path, rows: list[dict[str, Any]], *, drop_doc_number: bool = False) -> Path:
    wb = openpyxl.load_workbook(REGISTER, data_only=True)
    ws = wb["TFA"]
    for off, row in enumerate(rows):
        r = _TFA_FIRST_FREE_ROW + off
        ws.cell(row=r, column=_TFA["no"], value=70 + off)
        ws.cell(row=r, column=_TFA["issued"], value=row.get("issued", "26-09-20"))
        ws.cell(row=r, column=_TFA["sender"], value=row["sender"])
        ws.cell(row=r, column=_TFA["discipline"], value=row["discipline"])
        ws.cell(row=r, column=_TFA["seq"], value=row["seq"])
        ws.cell(row=r, column=_TFA["doc_number"], value=row.get("doc_number"))
        ws.cell(row=r, column=_TFA["title"], value=row["title"])
        ws.cell(row=r, column=_TFA["result"], value=row.get("result"))
    if drop_doc_number:
        ws.delete_cols(_TFA["doc_number"], 1)
    wb.save(dest)
    return dest


def _docs(client, auth, pid, include_orphaned=True):
    r = client.get(f"/api/projects/{pid}/documents", headers=auth("cm"),
                   params={"include_orphaned": include_orphaned, "page_size": 500})
    assert r.status_code == 200, r.text
    return r.json()["items"]


def _reviews(client, auth, pid, **params):
    r = client.get(f"/api/projects/{pid}/review-requests", headers=auth("cm"), params=params)
    assert r.status_code == 200, r.text
    return r.json()


def _readiness(client, auth, pid, act):
    r = client.get(f"/api/activities/{act}/readiness", headers=auth("cm"), params={"project_id": pid})
    assert r.status_code == 200, r.text
    return r.json()


def _resolve_for_doc(client, auth, pid, act, doc_id, decision, note):
    matches = [r for r in _reviews(client, auth, pid, kind="document_mapping", status="open")
               if r["activity_id"] == act and r["conflicting_sources"]["doc_id"] == doc_id]
    assert len(matches) == 1, f"{act}->{doc_id}: {matches}"
    r = client.post(f"/api/review-requests/{matches[0]['review_request_id']}/resolve",
                    headers=auth("cm"), json={"decision": decision, "note": note})
    assert r.status_code == 200, r.text
    return matches[0]


def _ids(client, auth, pid):
    return {d["sender_normalized"]: d["doc_id"] for d in _docs(client, auth, pid) if d["title"] == TITLE}


def _absorb(cfg):
    cfg["normalization"]["sender_aliases"]["중원엔지니어링"].append("중원E&C")


def _scenario(client, auth, user_ids, tmp_path, name, *, confirm_on, mutate=_absorb,
              drop_doc_number=False, week2_rows=(NEW,)):
    w1 = _register(tmp_path / "w1.xlsx", [OLD, NEW], drop_doc_number=drop_doc_number)
    w2 = _register(tmp_path / "w2.xlsx", list(week2_rows), drop_doc_number=drop_doc_number)
    pid = _new_project(client, auth, user_ids, name)
    upload(client, auth("contractor"), pid, SCHEDULE)
    _, first = upload(client, auth("cm"), pid, w1)
    assert first["status"] == "done", first
    assert first["result"]["identity_drift"] is None
    ids = _ids(client, auth, pid)
    assert set(ids) == {"중원엔지니어링", "중원E&C"}, ids
    _resolve_for_doc(client, auth, pid, ACT, ids[confirm_on], "approved", "이 도면을 근거로 삼는다")
    before = _readiness(client, auth, pid, ACT)
    if mutate is None:
        _, job = upload(client, auth("cm"), pid, w2)
    else:
        _, job = _upload_cfg(client, auth("cm"), pid, w2, _write_cfg(tmp_path / f"cfg-{confirm_on}", mutate))
    return {"pid": pid, "job": job, "survivor": ids["중원엔지니어링"], "vanished": ids["중원E&C"],
            "before": before, "after": _readiness(client, auth, pid, ACT)}


def _dump(tag, fx, client, auth):
    job = fx["job"]
    d = job["result"].get("identity_drift")
    print(f"\n### {tag}")
    print("  warnings:", [w["message"][:150] for w in (job.get("warnings") or [])])
    print("  drift:", None if d is None else {k: d[k] for k in ("moved", "merged", "lost_decisions")})
    print("  review_id:", job["result"].get("identity_drift_review_id"))
    print("  drawing_approval before/after:",
          fx["before"]["components"].get("drawing_approval"), "->", fx["after"]["components"].get("drawing_approval"))
    rv = _reviews(client, auth, fx["pid"], kind="document_identity_drift")
    for r in rv:
        print("  TITLE:", r["title"])
    return d, rv


def test_rp1_blocker_survivor(client, auth, user_ids, tmp_path):
    fx = _scenario(client, auth, user_ids, tmp_path, "RP1 사명변경 살아남는쪽", confirm_on="중원엔지니어링")
    d, rv = _dump("RP1 (blocker: 판단이 살아남는 행)", fx, client, auth)
    assert d is not None, "REGRESSION: identity_drift is None"
    assert len(rv) == 1, rv
    assert [x["cause"] for x in d["lost_decisions"]] == ["row_replaced"]
    assert d["lost_decisions"][0]["approval_flipped"] is True
    assert d["moved"] == [] and d["merged"] == []


def test_rp2_symmetric_vanished(client, auth, user_ids, tmp_path):
    fx = _scenario(client, auth, user_ids, tmp_path, "RP2 사명변경 사라지는쪽", confirm_on="중원E&C")
    d, rv = _dump("RP2 (대칭 짝: 판단이 사라지는 행)", fx, client, auth)
    assert d is not None
    assert len(rv) == 1, rv
    assert [x["cause"] for x in d["lost_decisions"]] == ["row_absorbed"]
    assert d["lost_decisions"][0]["new_doc_id"] == fx["survivor"]


# ─────────────────────────────────────────────────────────────────────────────
# RP3 — (다) "행-정체 전체 일치" 한정어의 역방향 확인
#   흡수가 일어난 **같은 주에 대장이 살아남는 행의 문서번호를 고치면** 짝짓기가 깨진다.
#   판단은 사라지는 쪽에 있다(= RP2 와 같은 사건). 발화하는가?
# ─────────────────────────────────────────────────────────────────────────────
OLD_A = {"sender": "중원", "discipline": "건축", "seq": SEQ, "result": "승인",
         "doc_number": "중원-HG-TFA-건축-26-210", "title": TITLE}
NEW_R = {"sender": "중원E&C", "discipline": "건축", "seq": SEQ, "result": "반려",
         "doc_number": "중원EC-HG-TFA-건축-26-211", "title": TITLE}
NEW_R_EDITED = dict(NEW_R, doc_number="중원EC-HG-TFA-건축-26-211-R1")


def _scenario2(client, auth, user_ids, tmp_path, name, week1_rows, week2_rows, confirm_sender):
    w1 = _register(tmp_path / "w1.xlsx", list(week1_rows))
    w2 = _register(tmp_path / "w2.xlsx", list(week2_rows))
    pid = _new_project(client, auth, user_ids, name)
    upload(client, auth("contractor"), pid, SCHEDULE)
    _, first = upload(client, auth("cm"), pid, w1)
    assert first["status"] == "done" and first["result"]["identity_drift"] is None
    ids = _ids(client, auth, pid)
    _resolve_for_doc(client, auth, pid, ACT, ids[confirm_sender], "approved", "근거로 삼는다")
    before = _readiness(client, auth, pid, ACT)
    _, job = _upload_cfg(client, auth("cm"), pid, w2, _write_cfg(tmp_path / "cfg", _absorb))
    return {"pid": pid, "job": job, "ids": ids, "before": before,
            "after": _readiness(client, auth, pid, ACT)}


def test_rp3a_absorbed_baseline_rejected_row_vanishes(client, auth, user_ids, tmp_path):
    """대조군: 문서번호를 안 고치면 흡수로 발화한다(판단은 사라지는 쪽 = 반려 행)."""
    fx = _scenario2(client, auth, user_ids, tmp_path, "RP3a 기준선", [OLD_A, NEW_R], [NEW_R], "중원E&C")
    d, rv = _dump("RP3a 대조군(문서번호 그대로)", fx, client, auth)
    assert d is not None and [x["cause"] for x in d["lost_decisions"]] == ["row_absorbed"]


def test_rp3b_absorbed_plus_same_week_doc_number_edit(client, auth, user_ids, tmp_path):
    """같은 주에 대장이 살아남는 행의 문서번호를 고쳤다. 행-정체 전체 일치가 깨진다."""
    fx = _scenario2(client, auth, user_ids, tmp_path, "RP3b 흡수+문서번호정정",
                    [OLD_A, NEW_R], [NEW_R_EDITED], "중원E&C")
    d, rv = _dump("RP3b 흡수 + 같은 주 문서번호 정정", fx, client, auth)
    print("  >>> RP3b drift is None?", d is None, " reviews:", len(rv))


# ─────────────────────────────────────────────────────────────────────────────
# FP — 내가 새로 만드는 오탐 후보
# ─────────────────────────────────────────────────────────────────────────────
def _fp(client, auth, user_ids, tmp_path, name, week1_row, week2_row):
    """대장만 바꾼다(config 무변). 같은 doc_id 가 유지되는지도 함께 본다."""
    w1 = _register(tmp_path / "w1.xlsx", [week1_row])
    w2 = _register(tmp_path / "w2.xlsx", [week2_row])
    pid = _new_project(client, auth, user_ids, name)
    upload(client, auth("contractor"), pid, SCHEDULE)
    _, first = upload(client, auth("cm"), pid, w1)
    assert first["status"] == "done", first
    ids = {d["doc_id"] for d in _docs(client, auth, pid) if d["title"].strip() == week1_row["title"].strip()}
    docs1 = [d for d in _docs(client, auth, pid) if d["seq_normalized"] == str(SEQ)]
    assert len(docs1) == 1, docs1
    doc_id = docs1[0]["doc_id"]
    _resolve_for_doc(client, auth, pid, ACT, doc_id, "approved", "근거로 삼는다")
    _, job = upload(client, auth("cm"), pid, w2)
    docs2 = [d for d in _docs(client, auth, pid) if d["seq_normalized"] == str(SEQ)]
    d = job["result"].get("identity_drift")
    print(f"\n### {name}")
    print("  doc_id 유지:", doc_id in {x['doc_id'] for x in docs2}, " 문서수:", len(docs2))
    print("  drift:", None if d is None else d["lost_decisions"])
    rv = _reviews(client, auth, pid, kind="document_identity_drift")
    for r in rv:
        print("  TITLE:", r["title"])
    return d, rv


BASE = {"sender": "중원", "discipline": "건축", "seq": "26-210", "result": "승인",
        "doc_number": "중원-HG-TFA-건축-26-210", "title": TITLE}


def test_fp1_seq_raw_format_change(client, auth, user_ids, tmp_path):
    """대장이 `번호` 표기를 '26-210' → '제26-210호' 로 바꿨다. seq_normalized 는 같다(숫자만)."""
    d, rv = _fp(client, auth, user_ids, tmp_path, "FP1 번호 표기 변경",
                BASE, dict(BASE, seq="제26-210호"))
    print("  >>> FP1 오탐인가:", d is not None and bool(d["lost_decisions"]))


def test_fp2_title_whitespace_only_change(client, auth, user_ids, tmp_path):
    """대장이 제목의 공백만 손댔다(두 칸). title_identity 는 같다 → doc_id 불변."""
    spaced = TITLE.replace(" - ", "  -  ")
    d, rv = _fp(client, auth, user_ids, tmp_path, "FP2 제목 공백만 변경",
                BASE, dict(BASE, title=spaced))
    print("  >>> FP2 오탐인가:", d is not None and bool(d["lost_decisions"]))


def test_fp3_sender_fullwidth_only(client, auth, user_ids, tmp_path):
    """발신을 같은 별칭 안에서 바꾼다('동부' → '(주)동부건설') — architect 가 P7 로 남긴 것의 내 재현."""
    base = dict(BASE, sender="동부", doc_number="동부-HG-TFA-건축-26-210")
    d, rv = _fp(client, auth, user_ids, tmp_path, "FP3 발신 표기 정정",
                base, dict(base, sender="(주)동부건설"))
    print("  >>> FP3 오탐인가:", d is not None and bool(d["lost_decisions"]))


def _blockers(rd):
    return [(b["component"], b.get("kind"), b.get("reason", "")[:60]) for b in rd["blockers"]]


def test_rp4_harm_direction_of_rp3(client, auth, user_ids, tmp_path):
    (tmp_path / "a").mkdir(); (tmp_path / "b").mkdir()
    a = _scenario2(client, auth, user_ids, tmp_path / "a", "RP4a", [OLD_A, NEW_R], [NEW_R], "중원E&C")
    b = _scenario2(client, auth, user_ids, tmp_path / "b", "RP4b", [OLD_A, NEW_R], [NEW_R_EDITED], "중원E&C")
    for tag, fx in (("RP4a 발화", a), ("RP4b 침묵", b)):
        print(f"\n### {tag}")
        print("  drawing_approval:", fx["before"]["components"]["drawing_approval"], "->",
              fx["after"]["components"]["drawing_approval"])
        print("  blockers before:", _blockers(fx["before"]))
        print("  blockers after :", _blockers(fx["after"]))
        print("  drift:", fx["job"]["result"].get("identity_drift") is not None)


def test_rp5_already_orphaned_counterpart(client, auth, user_ids, tmp_path):
    """(다)의 `was_orphaned` 제외 한정어. 세 번의 적재:
    1주 TFA:R(반려) + TFR:R'(행-정체 동일, 승인) → CM 이 TFA 쪽을 확정(차단 0.0)
    2주 TFR:R' 를 대장에서 뺀다 → Y 고아
    3주 대장이 R' 를 TFA 시트로 옮기고 옛 R 을 뺀다 → X 가 담은 행이 R' 로 바뀐다(승인)"""
    def reg(dest, tfa_rows, tfr_rows):
        wb = openpyxl.load_workbook(REGISTER, data_only=True)
        for off, row in enumerate(tfa_rows):
            r = _TFA_FIRST_FREE_ROW + off
            for k, col in (("no", "no"),):
                pass
            ws = wb["TFA"]
            ws.cell(row=r, column=_TFA["no"], value=70 + off)
            ws.cell(row=r, column=_TFA["issued"], value="26-09-20")
            ws.cell(row=r, column=_TFA["sender"], value=row["sender"])
            ws.cell(row=r, column=_TFA["discipline"], value=row["discipline"])
            ws.cell(row=r, column=_TFA["seq"], value=row["seq"])
            ws.cell(row=r, column=_TFA["doc_number"], value=row["doc_number"])
            ws.cell(row=r, column=_TFA["title"], value=row["title"])
            ws.cell(row=r, column=_TFA["result"], value=row["result"])
        tfr = wb["TFR"]
        _TFR = {"no": 1, "issued": 2, "sender": 3, "discipline": 4, "seq": 5, "doc_number": 6,
                "title": 7, "result": 8, "completed": 9}
        for off, row in enumerate(tfr_rows):
            r = 6 + off
            tfr.cell(row=r, column=_TFR["no"], value=80 + off)
            tfr.cell(row=r, column=_TFR["issued"], value="26-09-20")
            tfr.cell(row=r, column=_TFR["sender"], value=row["sender"])
            tfr.cell(row=r, column=_TFR["discipline"], value=row["discipline"])
            tfr.cell(row=r, column=_TFR["seq"], value=row["seq"])
            tfr.cell(row=r, column=_TFR["doc_number"], value=row["doc_number"])
            tfr.cell(row=r, column=_TFR["title"], value=row["title"])
            tfr.cell(row=r, column=_TFR["result"], value=row["result"])
        wb.save(dest)
        return dest

    R = {"sender": "중원", "discipline": "건축", "seq": SEQ, "result": "반려",
         "doc_number": "중원-HG-TFA-건축-26-210", "title": TITLE}
    R2 = dict(R, result="승인")     # 행-정체 4필드 동일, 처리결과만 다름
    w1 = reg(tmp_path / "w1.xlsx", [R], [R2])
    w2 = reg(tmp_path / "w2.xlsx", [R], [])
    w3 = reg(tmp_path / "w3.xlsx", [R2], [])

    pid = _new_project(client, auth, user_ids, "RP5 이미 고아였던 짝")
    upload(client, auth("contractor"), pid, SCHEDULE)
    _, j1 = upload(client, auth("cm"), pid, w1)
    tfa_docs = [d for d in _docs(client, auth, pid) if d["title"] == TITLE and d["doc_type"] == "TFA"]
    assert len(tfa_docs) == 1, tfa_docs
    X = tfa_docs[0]["doc_id"]
    _resolve_for_doc(client, auth, pid, ACT, X, "approved", "반려 도면임을 확인하고 차단한다")
    rd1 = _readiness(client, auth, pid, ACT)
    _, j2 = upload(client, auth("cm"), pid, w2)
    rd2 = _readiness(client, auth, pid, ACT)
    _, j3 = upload(client, auth("cm"), pid, w3)
    rd3 = _readiness(client, auth, pid, ACT)
    print("\n### RP5 (다)의 was_orphaned 제외")
    print("  drawing_approval:", rd1["components"]["drawing_approval"], "->",
          rd2["components"]["drawing_approval"], "->", rd3["components"]["drawing_approval"])
    for tag, j in (("w2", j2), ("w3", j3)):
        print(f"  {tag} drift:", j["result"].get("identity_drift"))
        print(f"  {tag} warn :", [w["message"][:80] for w in (j.get("warnings") or [])
                                  if "IDENTITY" in w["message"]])
    print("  drift reviews:", len(_reviews(client, auth, pid, kind="document_identity_drift")))
