"""문서 정체성 — 매칭 튜닝 내성 · 병합 탐지 · 식별 드리프트 (ADR 0009 / 계획 0003 §7 V2·V3·V4, 담당: qa).

재현된 결함(ADR 0009 §2): CM 이 검토 큐에서 A100 매핑을 확정하고 A400 매핑을 반려한 뒤,
`title_matching.normalize.strip_patterns` 에 `"승인요청"` 한 줄을 더하고 **같은 대장 파일을 한 바이트도
바꾸지 않은 채** 재업로드하면 문서 10건 중 6건이 새 `doc_id` 를 얻고, CM 의 확정은 고아 문서를 가리키게
되며, A400 에는 제목이 글자까지 같은 문서가 "반려된 것"과 "새로 검토해 달라는 것"으로 동시에 존재했다.
`drawing_approval` 은 1.0 → 0.5 로 떨어졌다. **그리고 이 전부가 조용했다** — job 은 `done`, 예외 없음,
`/startable` 200, 화면 정상.

**반증 — 결함이 있어도 통과하는 단언(계획 §7 V2·V3·V4, 실측으로 확인된 것들. 여기서는 쓰지 않거나,
쓰더라도 그것만으로 끝내지 않는다):**

- `job["status"] == "done"` — 결함 상태에서도 `done`.
- `job["result"]["mapping_count"] == 6` — 결함 상태에서도 6(새 doc_id 로 6건이 다시 만들어진다).
  게다가 **정상 상태에서는 4**다(사람이 판단한 2건은 `_drop_already_confirmed` 가 뺀다) — 이 숫자를
  단언하면 고쳐진 코드에서 오히려 실패한다. 계획 §7 V2 의 반증 목록이 이 값을 잘못 적었다.
- `len(documents(include_orphaned=False)) == 10` — 결함 상태에서도 10.
- `GET /startable` 200 — 결함 상태에서도 200.
- V2 에서 "뮤테이션이 실제로 적용됐다"(`title_normalized` 가 바뀌었다)를 확인하지 않으면 **뮤테이션이
  no-op 이어도 전부 초록**이 된다(config 키 오타 하나면 충분하다).
- V3 에서 `\\d+\\s*차` 뮤테이션만 두면 **탐지 코드가 없어도 통과한다** — 동결 이후 이 튜닝은 `doc_id` 를
  움직일 수 없으므로(실측 0/10) 충돌 자체가 일어나지 않기 때문이다. 그래서 (b) 강제 충돌 픽스처가
  반드시 필요하다. 또 `duplicate_doc_number` 경고에 기대면 안 된다 — 두 행의 문서번호를 다르게 두면
  그 경고는 뜨지 않는데 병합은 일어난다(**측정됨**).
- V4 에서 음성 대조군이 없으면 "고아가 생기면 드리프트"라는 틀린 구현(N1·N2)과 "항상 검토요청을
  만든다"는 구현(N3)이 그대로 통과한다.

**경고 단언은 `code` 필드가 아니라 메시지 접두사로 건다.** 대장 적재 잡의 경고는 전부
`DOCUMENT_REGISTER_WARNING` 으로 묶여 나가고(`services/api/jobs._warning`), 실제 code
(`DOCUMENT_IDENTITY_DRIFT`/`_COLLISION`)는 `RegisterWarning.__str__` 이 만드는 `"{code}[loc]: {detail}"`
의 접두사로 들어간다. 이 래핑은 기존 경고 전부의 계약이라 이번 사이클에서 바꾸지 않았다 — 그래서
`_has_warning()` 이 두 자리를 모두 본다.
"""
from __future__ import annotations

import copy
from collections.abc import Sequence
from pathlib import Path
from typing import Any

import openpyxl
import pytest
import yaml

from packages.core.settings import settings
from services.progress.config_loader import load_config

from .conftest import FIXTURES, add_member, upload

REGISTER = FIXTURES / "document_register.xlsx"
SCHEDULE = FIXTURES / "schedule.csv"
ACTIVITY_CONFIRM = "A100"    # 확정 대상 — 처리결과 APPROVED 인 TFA 가 매핑된다
ACTIVITY_REJECT = "A400"     # 반려 대상 — 같은 이유로 APPROVED 건을 고른다(누수가 보이도록)
FIXTURE_DOCUMENT_COUNT = 10

#: TFA 시트 열 번호(헤더 3행: No·문서발생일·발신·공종·번호·문서번호·회신요청일·제목·처리결과·처리완료일).
#: 열 위치를 상수로 박는 것은 **픽스처를 만드는 쪽**에서만 한다 — 파서는 헤더 별칭으로 찾아야 한다(ADR 0007 §2-5).
_TFA = {"no": 1, "issued": 2, "sender": 3, "discipline": 4, "seq": 5, "doc_number": 6,
        "reply_due": 7, "title": 8, "result": 9, "completed": 10}
_TFA_FIRST_FREE_ROW = 12     # 헤더 3행 + 데이터 8행


# ─────────────────────────────────────────────────────────────────────────────
# 공용 헬퍼
# ─────────────────────────────────────────────────────────────────────────────
def _new_project(client, auth, user_ids, name: str) -> str:
    r = client.post("/api/projects", headers=auth("admin"), json={"name": name})
    assert r.status_code == 201, r.text
    project_id = r.json()["project_id"]
    for role in ("contractor", "cm", "client"):
        add_member(client, auth("admin"), project_id, user_ids[role], role)
    return project_id


def _write_mutated_config(target: Path, mutate) -> Path:
    """실제 `config/document_register.yaml` 은 건드리지 않는다 — 값만 베껴 임시 디렉터리에 쓴다."""
    cfg = copy.deepcopy(load_config("document_register.yaml"))
    mutate(cfg)
    target.mkdir(parents=True, exist_ok=True)
    (target / "document_register.yaml").write_text(yaml.safe_dump(cfg, allow_unicode=True), encoding="utf-8")
    return target


def _upload_with_config(client, headers, project_id: str, path: Path, config_dir: Path) -> tuple[dict, dict]:
    """`settings.config_dir` 를 바꾼 채 업로드한다. 잡은 eager 라 이 호출 안에서 끝나므로 바로 되돌린다
    (monkeypatch 를 쓰지 않는 이유: 모듈 스코프 픽스처에서도 같은 방식이 필요하다)."""
    previous = settings.config_dir
    settings.config_dir = str(config_dir)
    try:
        return upload(client, headers, project_id, path)
    finally:
        settings.config_dir = previous


def _warning_messages(job: dict) -> list[str]:
    return [w["message"] for w in (job.get("warnings") or [])]


def _has_warning(job: dict, code: str) -> bool:
    return any(w.get("code") == code or str(w.get("message", "")).startswith(code)
               for w in (job.get("warnings") or []))


def _documents(client, auth, project_id: str, *, include_orphaned: bool = True) -> list[dict]:
    r = client.get(f"/api/projects/{project_id}/documents", headers=auth("cm"),
                   params={"include_orphaned": include_orphaned, "page_size": 500})
    assert r.status_code == 200, r.text
    return r.json()["items"]


def _document_detail(client, auth, project_id: str, doc_id: str) -> dict:
    r = client.get(f"/api/documents/{doc_id}", headers=auth("cm"), params={"project_id": project_id})
    assert r.status_code == 200, r.text
    return r.json()


def _reviews(client, auth, project_id: str, **params) -> list[dict]:
    r = client.get(f"/api/projects/{project_id}/review-requests", headers=auth("cm"), params=params)
    assert r.status_code == 200, r.text
    return r.json()


def _readiness(client, auth, project_id: str, activity_id: str) -> dict:
    r = client.get(f"/api/activities/{activity_id}/readiness", headers=auth("cm"),
                   params={"project_id": project_id})
    assert r.status_code == 200, r.text
    return r.json()


def _resolve_mapping_review(client, auth, project_id: str, activity_id: str, decision: str, note: str) -> dict:
    """CM 이 실제로 지나가는 유일한 경로 — 검토 큐(`POST /review-requests/{id}/resolve`)만 쓴다
    (`close_document_mapping_review`/`reject_document_mapping` 을 직접 부르지 않는다, test_15 와 같은 규약)."""
    open_reviews = _reviews(client, auth, project_id, kind="document_mapping", status="open")
    matches = [r for r in open_reviews if r["activity_id"] == activity_id]
    assert len(matches) == 1, f"{activity_id} 의 열린 매핑 검토요청이 정확히 1건이 아니다: {matches}"
    review = matches[0]
    r = client.post(f"/api/review-requests/{review['review_request_id']}/resolve", headers=auth("cm"),
                    json={"decision": decision, "note": note})
    assert r.status_code == 200, r.text
    return review


def _mapping(client, auth, project_id: str, doc_id: str, activity_id: str) -> list[dict]:
    return [m for m in _document_detail(client, auth, project_id, doc_id)["mappings"]
            if m["activity_id"] == activity_id]


def _register_with_extra_tfa_rows(dest: Path, rows: list[dict[str, Any]]) -> Path:
    """대장 픽스처의 TFA 시트 끝에 행을 덧붙인 사본. 원본 픽스처는 건드리지 않는다."""
    wb = openpyxl.load_workbook(REGISTER, data_only=True)
    ws = wb["TFA"]
    for offset, row in enumerate(rows):
        r = _TFA_FIRST_FREE_ROW + offset
        ws.cell(row=r, column=_TFA["no"], value=9 + offset)
        ws.cell(row=r, column=_TFA["issued"], value=row.get("issued", "26-09-20"))
        ws.cell(row=r, column=_TFA["sender"], value=row.get("sender", "동부"))
        ws.cell(row=r, column=_TFA["discipline"], value=row.get("discipline", "구조"))
        ws.cell(row=r, column=_TFA["seq"], value=row["seq"])
        ws.cell(row=r, column=_TFA["doc_number"], value=row["doc_number"])
        ws.cell(row=r, column=_TFA["title"], value=row["title"])
        ws.cell(row=r, column=_TFA["result"], value=row.get("result"))
    wb.save(dest)
    return dest


def _prepared_project(client, auth, user_ids, name: str) -> str:
    """정상 순서(공정표 → 대장)로 올린 프로젝트. 대장 적재가 6건의 매핑 + 6건의 열린 검토요청을 만든다."""
    project_id = _new_project(client, auth, user_ids, name)
    _, schedule_job = upload(client, auth("contractor"), project_id, SCHEDULE)
    assert schedule_job["status"] == "done", schedule_job
    _, register_job = upload(client, auth("cm"), project_id, REGISTER)
    assert register_job["status"] == "done", register_job
    assert register_job["result"]["identity_drift"] is None    # 첫 적재는 드리프트를 판정하지 않는다
    return project_id


# ═══════════════════════════════════════════════════════════════════════════
# V2 — 확정·반려가 매칭 튜닝을 견딘다
# ═══════════════════════════════════════════════════════════════════════════
@pytest.fixture(scope="module")
def matching_tuned(client, auth, user_ids, tmp_path_factory) -> dict[str, Any]:
    """재현 절차 그대로: 검토 큐로만 A100 확정 · A400 반려 → 매칭 config 뮤테이션 → **같은 대장** 재업로드."""
    project_id = _prepared_project(client, auth, user_ids, "V2 매칭 튜닝 내성")
    confirm_review = _resolve_mapping_review(client, auth, project_id, ACTIVITY_CONFIRM,
                                             "approved", "대장 확인 결과 승인 확인")
    reject_review = _resolve_mapping_review(client, auth, project_id, ACTIVITY_REJECT,
                                            "rejected", "이 문서는 이 작업과 무관하다")
    before_docs = {d["doc_id"]: d for d in _documents(client, auth, project_id)}
    before_readiness = _readiness(client, auth, project_id, ACTIVITY_CONFIRM)

    config_dir = _write_mutated_config(
        tmp_path_factory.mktemp("v2-config"),
        lambda cfg: cfg["title_matching"]["normalize"]["strip_patterns"].append("승인요청"))
    _, job = _upload_with_config(client, auth("cm"), project_id, REGISTER, config_dir)

    return {"project_id": project_id, "job": job,
            "confirmed_doc_id": confirm_review["conflicting_sources"]["doc_id"],
            "rejected_doc_id": reject_review["conflicting_sources"]["doc_id"],
            "before_docs": before_docs, "before_readiness": before_readiness,
            "after_docs": {d["doc_id"]: d for d in _documents(client, auth, project_id)}}


def test_v2_matching_tuning_creates_no_orphans_and_no_drift(matching_tuned) -> None:
    """계약의 관측 가능한 얼굴: 같은 대장을 다시 올렸는데 아무 문서도 새로 생기지 않고 고아도 없다.

    결함 상태의 같은 자리는 `created=6, orphaned=6` 이었다(ADR 0009 §2 실측).
    """
    result = matching_tuned["job"]["result"]
    assert (result["created"], result["orphaned"], result["updated"]) == (0, 0, FIXTURE_DOCUMENT_COUNT)
    assert result["identity_drift"] is None
    assert result["identity_drift_moved"] == 0 and result["identity_drift_lost_decisions"] == 0
    assert not _has_warning(matching_tuned["job"], "DOCUMENT_IDENTITY_DRIFT")
    assert not _has_warning(matching_tuned["job"], "document_possibly_renamed")
    assert set(matching_tuned["after_docs"]) == set(matching_tuned["before_docs"])


def test_v2_mutation_actually_applied_to_matching_normalization(matching_tuned) -> None:
    """**가짜 초록 방지(계획 §7 V2 5번).** 뮤테이션이 no-op 이면 위·아래 단언은 아무것도 증명하지 않는다.
    `title_normalized`(대조용)는 실제로 6건 바뀌어 있어야 한다 — 바뀐 것은 대조 정규화뿐이고 정체성은
    그대로라는 것이 ADR 0009 의 계약이다."""
    changed = [doc_id for doc_id, after in matching_tuned["after_docs"].items()
               if after["title_normalized"] != matching_tuned["before_docs"][doc_id]["title_normalized"]]
    assert len(changed) == 6, changed
    # 같은 문서의 식별용 제목은 한 글자도 움직이지 않았다.
    assert all(matching_tuned["after_docs"][doc_id]["title_identity"]
               == matching_tuned["before_docs"][doc_id]["title_identity"] for doc_id in changed)


def test_v2_confirmed_mapping_survives_the_tuning(client, auth, user_ids, matching_tuned) -> None:
    """CM 이 확정한 매핑이 그대로 남고, 그 문서는 고아가 되지 않는다(결함 상태에서는 둘 다 반대였다)."""
    project_id, doc_id = matching_tuned["project_id"], matching_tuned["confirmed_doc_id"]
    detail = _document_detail(client, auth, project_id, doc_id)
    assert detail["document"]["is_orphaned"] is False

    mappings = [m for m in detail["mappings"] if m["activity_id"] == ACTIVITY_CONFIRM]
    assert len(mappings) == 1, mappings          # 새 doc_id 로 후보가 다시 열리지 않았다
    assert mappings[0]["needs_review"] is False and mappings[0]["reviewed_by"] == user_ids["cm"]
    assert not [r for r in _reviews(client, auth, project_id, kind="document_mapping", status="open")
                if r["activity_id"] == ACTIVITY_CONFIRM]


def test_v2_rejection_stays_permanent(client, auth, matching_tuned) -> None:
    """ADR 0007 §4-2 규칙 6 ⑥(반려의 영구성). 결함 상태에서는 제목이 글자까지 같은 문서가 "반려된 것"과
    "새로 검토해 달라는 것"으로 **동시에** 존재했다."""
    project_id, doc_id = matching_tuned["project_id"], matching_tuned["rejected_doc_id"]
    mappings = _mapping(client, auth, project_id, doc_id, ACTIVITY_REJECT)
    assert len(mappings) == 1, mappings
    assert mappings[0]["evidence"]["extra"]["mapping_review_decision"] == "rejected"

    reject_reviews = [r for r in _reviews(client, auth, project_id, kind="document_mapping")
                      if r["activity_id"] == ACTIVITY_REJECT]
    assert [r["status"] for r in reject_reviews] == ["rejected"]


def test_v2_drawing_approval_readiness_is_unchanged(client, auth, matching_tuned) -> None:
    """착수 가능 판단이 config 한 줄에 움직이지 않는다. 결함 상태에서는 `drawing_approval` 1.0 → 0.5,
    총점 1.0 → 0.925 였다(그리고 blocker 가 새 doc_id 를 가리켰다)."""
    after = _readiness(client, auth, matching_tuned["project_id"], ACTIVITY_CONFIRM)
    before = matching_tuned["before_readiness"]
    assert after["components"]["drawing_approval"] == before["components"]["drawing_approval"] == 1.0
    assert after["score"] == before["score"]
    assert not [b for b in after["blockers"] if b["component"] == "drawing_approval"]


# ═══════════════════════════════════════════════════════════════════════════
# V3 — 병합(공격적 실패) 탐지
# ═══════════════════════════════════════════════════════════════════════════
@pytest.fixture(scope="module")
def revision_register(tmp_path_factory) -> Path:
    """반려 후 재제출이라 같은 `번호` 아래 1차·2차 두 행이 있는 대장(현장에서 흔하다).
    문서번호는 서로 다르게 둔다 — `duplicate_doc_number` 경고에 기대지 않기 위해서다."""
    return _register_with_extra_tfa_rows(tmp_path_factory.mktemp("v3") / "revisions.xlsx", [
        {"seq": 26060, "doc_number": "동부-HG-TFA-구조-26-060", "result": "반려",
         "title": "시공상세도 승인요청 - 1F 기둥 배근도 (Z1) 1차"},
        {"seq": 26060, "doc_number": "동부-HG-TFA-구조-26-061", "result": "승인",
         "title": "시공상세도 승인요청 - 1F 기둥 배근도 (Z1) 2차"},
    ])


def test_v3a_revision_tuning_no_longer_merges_documents(client, auth, user_ids, tmp_path, revision_register) -> None:
    """(a) `strip_patterns += r"\\d+\\s*차"` — 매칭 관점에서는 옳은 튜닝이다(`discriminative_tokens` 에
    이미 `revision` 규칙이 있다). ADR 0009 이전에는 이 한 줄이 1차(반려)를 2차(승인) 뒤로 지웠다.

    **이 단언만으로는 탐지 코드의 존재를 증명하지 못한다**(계획 §7 V3 반증) — 지금은 충돌 자체가
    일어나지 않기 때문이다. 그것을 (b)가 맡는다.
    """
    project_id = _new_project(client, auth, user_ids, "V3a 차수 튜닝")
    upload(client, auth("contractor"), project_id, SCHEDULE)
    _, first = upload(client, auth("cm"), project_id, revision_register)
    assert first["result"]["document_count"] == FIXTURE_DOCUMENT_COUNT + 2

    config_dir = _write_mutated_config(
        tmp_path / "cfg", lambda cfg: cfg["title_matching"]["normalize"]["strip_patterns"].append(r"\d+\s*차"))
    _, job = _upload_with_config(client, auth("cm"), project_id, revision_register, config_dir)

    docs = _documents(client, auth, project_id)
    assert job["result"]["document_count"] == len(docs) == FIXTURE_DOCUMENT_COUNT + 2
    assert len({d["doc_id"] for d in docs}) == len(docs)           # 충돌 0건
    assert (job["result"]["created"], job["result"]["orphaned"]) == (0, 0)
    assert not _has_warning(job, "DOCUMENT_IDENTITY_COLLISION")
    # 1차(반려)가 2차(승인) 뒤로 사라지지 않았다 — 승인 상태가 두 행에 따로 남아 있다.
    revisions = {d["doc_number"]: d["approval_status"] for d in docs if "(Z1) 1차" in d["title"] or "(Z1) 2차" in d["title"]}
    assert revisions == {"동부-HG-TFA-구조-26-060": "REJECTED", "동부-HG-TFA-구조-26-061": "APPROVED"}


def test_v3b_forced_collision_is_reported_instead_of_silently_dropping_a_row(
    client, auth, user_ids, tmp_path,
) -> None:
    """(b) 강제 충돌 — `doc_type`·`발신`·`번호`·제목 원문이 전부 같고 **문서번호만 다른** 두 행.

    병합은 되돌릴 수 없는 실패다(ADR 0009 §3 (나)): upsert 루프에서 마지막 행이 이기므로 앞 행이
    사라지고, 살아남은 행의 `approval_status` 가 `drawing_approval` 논리곱의 입력이 된다.
    `document_count`(12)와 실제 행 수(11)가 어긋나는데 `created`+`updated` 합은 맞아떨어져 **산술로도
    드러나지 않는다.** 덮어쓰기 동작 자체는 유지한다(대장이 정본) — 다만 더 이상 조용하지 않다.
    """
    project_id = _new_project(client, auth, user_ids, "V3b 강제 충돌")
    upload(client, auth("contractor"), project_id, SCHEDULE)
    collided = _register_with_extra_tfa_rows(tmp_path / "collision.xlsx", [
        {"seq": 26070, "doc_number": "동부-HG-TFA-구조-26-070", "result": "반려",
         "title": "시공상세도 승인요청 - 3F 기둥 배근도 (Z9)"},
        {"seq": 26070, "doc_number": "동부-HG-TFA-구조-26-071", "result": "승인",
         "title": "시공상세도 승인요청 - 3F 기둥 배근도 (Z9)"},
    ])
    _, job = upload(client, auth("cm"), project_id, collided)

    assert _has_warning(job, "DOCUMENT_IDENTITY_COLLISION"), _warning_messages(job)
    # 문서번호가 다르므로 `duplicate_doc_number` 는 뜨지 않는다 — 그 경고에 기대면 병합을 놓친다(측정됨).
    assert not _has_warning(job, "duplicate_doc_number")

    drift = job["result"]["identity_drift"]
    assert drift is not None and len(drift["merged"]) == 1
    assert drift["merged"][0]["titles"] == ["시공상세도 승인요청 - 3F 기둥 배근도 (Z9)"] * 2
    assert job["result"]["identity_drift_merged"] == 1

    docs = _documents(client, auth, project_id)
    assert job["result"]["document_count"] == FIXTURE_DOCUMENT_COUNT + 2      # 대장 행은 12
    assert len(docs) == FIXTURE_DOCUMENT_COUNT + 1                            # 실제 남은 행은 11
    survivors = [d for d in docs if d["title"] == "시공상세도 승인요청 - 3F 기둥 배근도 (Z9)"]
    assert [(d["doc_number"], d["approval_status"]) for d in survivors] == [("동부-HG-TFA-구조-26-071", "APPROVED")]

    # 사람의 판단이 걸려 있지 않으므로 검토요청까지 만들지는 않는다(ADR 0009 §5-2 — 큐 오염 방지).
    assert job["result"]["identity_drift_review_id"] is None
    assert _reviews(client, auth, project_id, kind="document_identity_drift") == []


# ═══════════════════════════════════════════════════════════════════════════
# V4 — 식별 표면 드리프트 탐지 + 음성 대조군
# ═══════════════════════════════════════════════════════════════════════════
def _rename_sender_standard_name(cfg: dict[str, Any]) -> None:
    """동결할 수 없는 식별 표면(ADR 0009 §5-1) — 새 협력사가 들어오면 반드시 만지는 표다. 실측 7/10 이동."""
    aliases = cfg["normalization"]["sender_aliases"]
    aliases["동부건설(주)"] = aliases.pop("동부건설")


@pytest.fixture(scope="module")
def sender_alias_drift(client, auth, user_ids, tmp_path_factory) -> dict[str, Any]:
    project_id = _prepared_project(client, auth, user_ids, "V4 식별 표면 드리프트")
    review = _resolve_mapping_review(client, auth, project_id, ACTIVITY_CONFIRM, "approved", "확정")
    confirmed_doc_id = review["conflicting_sources"]["doc_id"]
    mappings_before = _mapping(client, auth, project_id, confirmed_doc_id, ACTIVITY_CONFIRM)
    confirmed_before = _document_detail(client, auth, project_id, confirmed_doc_id)["document"]

    config_dir = _write_mutated_config(tmp_path_factory.mktemp("v4-config"), _rename_sender_standard_name)
    _, job = _upload_with_config(client, auth("cm"), project_id, REGISTER, config_dir)
    # 확정이 걸려 있던 대장 행이 지금 어느 `doc_id` 아래에 있는가 — **드리프트 보고서를 보지 않고**
    # 대장 원문(문서번호·제목)으로 찾는다. 보고서에서 읽어 오면 "보고서가 자기 자신과 같다"만 남는다.
    moved_to = [d["doc_id"] for d in _documents(client, auth, project_id)
                if d["doc_number"] == confirmed_before["doc_number"] and d["title"] == confirmed_before["title"]
                and d["doc_id"] != confirmed_doc_id]
    assert len(moved_to) == 1, moved_to
    return {"project_id": project_id, "job": job, "confirmed_doc_id": confirmed_doc_id,
            "moved_to_doc_id": moved_to[0], "mappings_before": mappings_before}


def test_v4_identity_surface_change_raises_a_drift_warning(sender_alias_drift) -> None:
    """대장 원문은 그대로인데 `doc_id` 가 이동했다 — 우리 쪽 식별 규칙이 바뀐 것이다.
    이때 `document_possibly_renamed`(제목이 바뀌었다는 뜻)는 **뜨지 않아야 한다**: 제목은 한 글자도
    바뀌지 않았고, 그 문구는 사실과 다르다(ADR 0009 §Deferred 2)."""
    job = sender_alias_drift["job"]
    assert _has_warning(job, "DOCUMENT_IDENTITY_DRIFT"), _warning_messages(job)
    assert not _has_warning(job, "document_possibly_renamed")
    drift = job["result"]["identity_drift"]
    assert drift is not None and len(drift["moved"]) == 7
    assert drift["previous_fingerprint"] != drift["current_fingerprint"]   # config 가 바뀌었다
    assert job["result"]["identity_drift_moved"] == 7
    moved_titles = {m["title"] for m in drift["moved"]}
    assert len(moved_titles) == 7   # 이동 쌍은 제목 원문이 **같은** 짝이다(1:1)


def test_v4_lost_human_decision_opens_exactly_one_cm_review(client, auth, sender_alias_drift) -> None:
    """잃어버린 사람의 판단이 실제로 있을 때만, **적재당 1건** CM 큐에 올린다(ADR 0009 §5-2·§5-3).

    8차 리뷰가 겪은 실패의 반대편 대비다 — 그때는 `needs_review=True` 매핑이 쌓여도 검토요청을 만드는
    코드가 없어 CM 큐가 영원히 비어 있었고 어떤 테스트도 실패하지 않았다.
    """
    project_id = sender_alias_drift["project_id"]
    reviews = _reviews(client, auth, project_id, kind="document_identity_drift")
    assert len(reviews) == 1, reviews
    review = reviews[0]
    assert review["status"] == "open" and review["assignee_role"] == "cm" and review["activity_id"] is None
    assert review["review_request_id"] == sender_alias_drift["job"]["result"]["identity_drift_review_id"]
    # 경위와 그 경위가 아는 값 넷이 항목마다 실린다(ADR 0009 §5-2 (마) 항목 계약, 개정 2).
    # `row_moved` 는 옮겨간 곳이 **있으므로** `new_doc_id` 가 차 있고, 행-정체는 손대지 않았으므로
    # `changed_fields` 는 비고, 현재 행이 없으므로 `approval_flipped` 는 언제나 False 다.
    assert review["conflicting_sources"]["lost_decisions"] == [
        {"activity_id": ACTIVITY_CONFIRM, "doc_id": sender_alias_drift["confirmed_doc_id"],
         "decision": "confirmed", "cause": "row_moved",
         "new_doc_id": sender_alias_drift["moved_to_doc_id"],
         "changed_fields": [], "approval_flipped": False}]
    assert review["confidence"] == 1.0                       # 판정이 아니라 관측이다
    assert review["evidence"]["method"] == "identity_drift_detection"
    assert "복구되지 않습니다" in review["title"]              # 매핑 복구를 약속하지 않는다


def test_v4_resolving_the_drift_review_changes_no_mapping(client, auth, user_ids, sender_alias_drift) -> None:
    """§5-3 — 이 kind 는 **확인 전용**이다. 해소는 `status`/`resolution_note`/`resolved_by` 만 남기고
    매핑을 되살리지 않는다(시스템이 사람의 확정을 복원하는 것은 ADR 0001 불변식과 충돌한다)."""
    project_id, doc_id = sender_alias_drift["project_id"], sender_alias_drift["confirmed_doc_id"]
    review = _reviews(client, auth, project_id, kind="document_identity_drift")[0]
    documents_before = _documents(client, auth, project_id)

    r = client.post(f"/api/review-requests/{review['review_request_id']}/resolve", headers=auth("cm"),
                    json={"decision": "approved", "note": "config 되돌림 없이 그대로 두기로 확인"})
    assert r.status_code == 200, r.text
    assert r.json()["status"] == "approved" and r.json()["resolved_by"] == user_ids["cm"]

    assert _documents(client, auth, project_id) == documents_before
    assert _mapping(client, auth, project_id, doc_id, ACTIVITY_CONFIRM) == sender_alias_drift["mappings_before"]


def test_v4_sheet_rename_is_detected_even_though_nothing_is_orphaned(
    client, auth, user_ids, tmp_path,
) -> None:
    """**음성이 아니라 사각지대**(계획 §1-a 블라인드 스팟 1): 사용자가 엑셀에서 시트명을 바꾸면 config 는
    한 글자도 안 바뀌었는데 `doc_type` 이 바뀌어 `doc_id` 가 움직인다.

    이때 옛 행은 **고아가 되지도 않는다** — 재업로드 규칙(ADR 0007 §2-2 규칙 2)이 "이번 업로드에 등장한
    doc_type" 만 고아 처리하는데, 옛 `TFA` 는 이번 업로드에 등장하지 않기 때문이다(실측: `orphaned=0`).
    드리프트 후보를 "고아"로 좁힌 구현은 이 경로 전체를 조용히 놓친다. 지문도 그대로다(config 가 안
    바뀌었으므로) — **지문 변화는 판정 조건이 아니라 보고 값이다.**
    """
    project_id = _prepared_project(client, auth, user_ids, "V4 시트명 변경")
    _resolve_mapping_review(client, auth, project_id, ACTIVITY_CONFIRM, "approved", "확정")

    wb = openpyxl.load_workbook(REGISTER, data_only=True)
    wb["TFA"].title = "자료제출"          # sheet_doc_types 에서 TFR 로 걸린다
    renamed = tmp_path / "sheet_renamed.xlsx"
    wb.save(renamed)
    _, job = upload(client, auth("cm"), project_id, renamed)

    assert job["result"]["orphaned"] == 0, "이 경로는 고아를 만들지 않는다 — 그것이 사각지대의 핵심이다"
    assert _has_warning(job, "DOCUMENT_IDENTITY_DRIFT"), _warning_messages(job)
    drift = job["result"]["identity_drift"]
    assert drift is not None and len(drift["moved"]) == 8
    assert drift["previous_fingerprint"] == drift["current_fingerprint"], "config 는 바뀌지 않았다"
    assert len(_reviews(client, auth, project_id, kind="document_identity_drift")) == 1


# ── 음성 대조군 ───────────────────────────────────────────────────────────────
def test_v4_n1_real_deletion_orphans_without_reporting_drift(client, auth, user_ids, tmp_path) -> None:
    """N1 — 대장에서 문서를 **진짜로 지웠다**. 고아는 생기지만 드리프트가 아니다(우리 규칙은 그대로다).
    이 대조군이 없으면 "고아가 생기면 드리프트"라는 틀린 구현이 그대로 통과한다."""
    project_id = _prepared_project(client, auth, user_ids, "V4 N1 삭제")
    _resolve_mapping_review(client, auth, project_id, ACTIVITY_CONFIRM, "approved", "확정")

    wb = openpyxl.load_workbook(REGISTER, data_only=True)
    wb["TFA"].delete_rows(11, 1)          # 마지막 TFA 행(통신 배관 경로 검토요청)을 지운다
    reduced = tmp_path / "deleted_row.xlsx"
    wb.save(reduced)
    _, job = upload(client, auth("cm"), project_id, reduced)

    assert job["result"]["orphaned"] == 1 and job["result"]["created"] == 0
    assert job["result"]["identity_drift"] is None
    assert not _has_warning(job, "DOCUMENT_IDENTITY_DRIFT")
    assert _reviews(client, auth, project_id, kind="document_identity_drift") == []


def test_v4_n2_real_retitle_is_a_rename_not_a_drift(client, auth, user_ids, tmp_path) -> None:
    """N2 — 사람이 대장에서 **제목을 진짜로 고쳤다**. 정체성이 흔들릴 수 있는 정상 상황이므로
    `document_possibly_renamed` 로 남기고 드리프트로 부르지 않는다(ADR 0009 §5-2 표 1행)."""
    project_id = _prepared_project(client, auth, user_ids, "V4 N2 제목 수정")
    _resolve_mapping_review(client, auth, project_id, ACTIVITY_CONFIRM, "approved", "확정")

    wb = openpyxl.load_workbook(REGISTER, data_only=True)
    wb["TFA"].cell(row=4, column=_TFA["title"], value="시공상세도 승인요청 - 1F 기둥 배근도 (Z1) 개정본")
    retitled = tmp_path / "retitled.xlsx"
    wb.save(retitled)
    _, job = upload(client, auth("cm"), project_id, retitled)

    assert (job["result"]["created"], job["result"]["orphaned"]) == (1, 1)
    assert _has_warning(job, "document_possibly_renamed"), _warning_messages(job)
    assert not _has_warning(job, "DOCUMENT_IDENTITY_DRIFT")
    assert job["result"]["identity_drift"] is None
    assert _reviews(client, auth, project_id, kind="document_identity_drift") == []


def test_v4_n3_drift_without_human_decisions_warns_but_opens_no_review(
    client, auth, user_ids, tmp_path,
) -> None:
    """N3 — 사람의 판단이 하나도 없는 프로젝트. 경고는 뜨지만 검토요청은 만들지 않는다(ADR 0009 §5-2).

    이 대조군이 없으면 "항상 검토요청을 만든다"는 구현이 통과하고, 그러면 새 협력사를 추가한 주마다
    CM 큐가 오염되는 설계를 테스트가 승인해 버린다 — 그 끝은 운영자가 **탐지를 끄는 것**이다.
    """
    project_id = _prepared_project(client, auth, user_ids, "V4 N3 판단 없음")
    assert not _reviews(client, auth, project_id, kind="document_mapping", status="approved")

    config_dir = _write_mutated_config(tmp_path / "cfg", _rename_sender_standard_name)
    _, job = _upload_with_config(client, auth("cm"), project_id, REGISTER, config_dir)

    assert _has_warning(job, "DOCUMENT_IDENTITY_DRIFT"), _warning_messages(job)
    drift = job["result"]["identity_drift"]
    assert drift is not None and len(drift["moved"]) == 7 and drift["lost_decisions"] == []
    assert job["result"]["identity_drift_review_id"] is None
    assert _reviews(client, auth, project_id, kind="document_identity_drift") == []


# ═══════════════════════════════════════════════════════════════════════════
# V5 — 병합이 **사람의 판단을 오염시킨다**(ADR 0009 §3 (나))
#
# V3b 는 병합을 탐지하지만 그 충돌에 **사람 판단을 걸지 않았다** — 그래서 "병합은 절대로 CM 큐에
# 올라가지 않는다"는 구멍이 테스트에 보이지 않았다. 실행으로 확인된 모습은 그보다 나빴다:
# CM 이 "반려된 도면"임을 확인하고 A300 매핑을 확정해 차단해 둔 상태(`drawing_approval` 0.0)에서
# `sender_aliases` 별칭표 통합 한 줄이 그 문서의 승인 상태를 REJECTED → APPROVED 로 뒤집어
# `drawing_approval` 0.0 → 1.0, **미승인 도면 위에서 착수 가능**이 떴고 검토요청은 생기지 않았다.
# ADR 0009 §3 이 스스로 최악이라 적은 문장 그대로다.
#
# 뒤집힘 자체는 의도된 동작이다 — 살아남은 행은 대장의 다른 행이고 **대장이 정본**이다(ADR 0007 §1
# 규칙 1). 이번 사이클이 바꾼 것은 그것이 더 이상 **조용하지 않다**는 점이므로, 아래 양성 테스트는
# 뒤집힘과 CM 큐를 **함께** 고정한다. 하나만 고정하면 다른 하나가 사라져도 초록이다.
# ═══════════════════════════════════════════════════════════════════════════
ACTIVITY_MERGE = "A300"                                       # 1F 덕트 설치 — 아래 두 행이 모두 이 Activity 의 후보가 된다
_MERGE_TITLE = "시공상세도 승인요청 - 1F 덕트 설치 상세도 (Z1)"   # 두 행의 제목 원문이 **글자 그대로** 같아야 병합이 성립한다
_MERGE_SEQ = 26080                                            # 같은 `번호` — doc_id 재료 셋 중 둘이 이미 같다
#: 대장 행 순서 = upsert 승자 순서. 뒤 행(승인)이 앞 행(반려)을 덮어쓴다(ADR 0009 §3 (나)).
_MERGE_ROWS = [
    {"sender": "동부", "discipline": "기계", "seq": _MERGE_SEQ, "result": "반려",
     "doc_number": "동부-HG-TFA-기계-26-080", "title": _MERGE_TITLE},
    {"sender": "동부이앤씨", "discipline": "기계", "seq": _MERGE_SEQ, "result": "승인",
     "doc_number": "동부이앤씨-HG-TFA-기계-26-081", "title": _MERGE_TITLE},
]
#: 충돌 상시화 대조군(N1b·N2 용): 두 행이 **첫 적재부터** 같은 doc_id 로 수렴한다(V3b 와 같은 모양).
_STANDING_COLLISION_ROWS = [
    {"sender": "동부", "discipline": "구조", "seq": 26090, "result": "반려",
     "doc_number": "동부-HG-TFA-구조-26-090", "title": "시공상세도 승인요청 - 3F 기둥 배근도 (Z9)"},
    {"sender": "동부", "discipline": "구조", "seq": 26090, "result": "승인",
     "doc_number": "동부-HG-TFA-구조-26-091", "title": "시공상세도 승인요청 - 3F 기둥 배근도 (Z9)"},
]


def _merge_sender_aliases(cfg: dict[str, Any]) -> None:
    """별칭표 **통합** — 새 법인명을 기존 표준명 아래로 넣는 운영상 정상 변경(ADR 0009 §5-1: 동결할 수
    없는 식별 표면). 이 한 줄이 서로 다른 두 대장 행을 같은 `doc_id` 로 붕괴시킨다."""
    cfg["normalization"]["sender_aliases"]["동부건설"].append("동부이앤씨")


def _resolve_mapping_review_for_doc(client, auth, project_id: str, activity_id: str, doc_id: str,
                                    decision: str, note: str) -> dict:
    """`_resolve_mapping_review` 의 doc_id 판(A300 에는 후보가 여럿이라 Activity 만으로는 못 고른다).
    CM 이 지나가는 경로는 여기서도 검토 큐 하나뿐이다."""
    matches = [r for r in _reviews(client, auth, project_id, kind="document_mapping", status="open")
               if r["activity_id"] == activity_id and r["conflicting_sources"]["doc_id"] == doc_id]
    assert len(matches) == 1, f"{activity_id}→{doc_id} 의 열린 매핑 검토요청이 정확히 1건이 아니다: {matches}"
    r = client.post(f"/api/review-requests/{matches[0]['review_request_id']}/resolve", headers=auth("cm"),
                    json={"decision": decision, "note": note})
    assert r.status_code == 200, r.text
    return matches[0]


def _duct_doc_ids(client, auth, project_id: str) -> dict[str, str]:
    """제목이 같은 두 덕트 행을 `sender_normalized` 로 가른다 — 병합 전에는 그것만이 둘을 구별한다."""
    return {d["sender_normalized"]: d["doc_id"]
            for d in _documents(client, auth, project_id) if d["title"] == _MERGE_TITLE}


def _drift_reviews(client, auth, project_id: str) -> list[dict]:
    return _reviews(client, auth, project_id, kind="document_identity_drift")


def _drawing_blockers(readiness: dict) -> list[dict]:
    return [b for b in readiness["blockers"] if b["component"] == "drawing_approval"]


@pytest.fixture(scope="module")
def merge_register(tmp_path_factory) -> Path:
    """제목·번호·종류가 같고 **발신 표기만 다른** 두 행이 있는 대장. 별칭표가 둘을 갈라 두는 동안에는
    서로 다른 문서다(첫 적재에서 충돌 0건)."""
    return _register_with_extra_tfa_rows(tmp_path_factory.mktemp("v5") / "merge.xlsx", _MERGE_ROWS)


# ── 양성 1: 확정이 **살아남는 행**에 있고 병합이 그 행을 덮어쓴다 ─────────────────
@pytest.fixture(scope="module")
def merge_overwritten(client, auth, user_ids, tmp_path_factory, merge_register) -> dict[str, Any]:
    """CM 이 **반려된 도면**(동부, REJECTED)의 매핑을 확정해 A300 을 차단해 둔 뒤 별칭표를 통합한다.
    병합 후 그 `doc_id` 는 살아남지만 담긴 대장 행은 **승인된 다른 행**(동부이앤씨)으로 바뀐다."""
    project_id = _new_project(client, auth, user_ids, "V5 병합 — 살아남는 행")
    upload(client, auth("contractor"), project_id, SCHEDULE)
    _, first = upload(client, auth("cm"), project_id, merge_register)
    assert first["status"] == "done" and first["result"]["identity_drift"] is None   # 아직 충돌이 없다

    doc_ids = _duct_doc_ids(client, auth, project_id)
    survivor = doc_ids["동부건설"]        # 별칭 통합 뒤에도 이 doc_id 가 그대로 남는다(재료가 바뀌지 않는다)
    _resolve_mapping_review_for_doc(client, auth, project_id, ACTIVITY_MERGE, survivor,
                                    "approved", "반려된 도면임을 확인 — 이 작업의 도면 근거로 삼는다")
    before_document = _document_detail(client, auth, project_id, survivor)["document"]
    before_readiness = _readiness(client, auth, project_id, ACTIVITY_MERGE)
    before_mapping = _mapping(client, auth, project_id, survivor, ACTIVITY_MERGE)

    config_dir = _write_mutated_config(tmp_path_factory.mktemp("v5-overwritten"), _merge_sender_aliases)
    _, job = _upload_with_config(client, auth("cm"), project_id, merge_register, config_dir)
    return {"project_id": project_id, "job": job, "survivor": survivor, "absorbed": doc_ids["동부이앤씨"],
            "before_document": before_document, "before_readiness": before_readiness,
            "before_mapping": before_mapping, "config_dir": config_dir, "register": merge_register}


def test_v5_merge_overwritten_puts_the_polluted_decision_on_the_cm_queue(client, auth, merge_overwritten) -> None:
    """병합이 **살아남은 행의 내용**을 갈아치웠다 — 판단이 사라진 게 아니라 판단의 **대상**이 바뀌었다.
    이동(`moved`)이 0건이라 옛 판정(고아 ↔ 신규 짝짓기)으로는 잡히지 않는 경로이고, 그래서 이 사건이
    사이클 끝까지 CM 큐에 닿지 못했다."""
    project_id, survivor = merge_overwritten["project_id"], merge_overwritten["survivor"]
    result = merge_overwritten["job"]["result"]
    assert (result["identity_drift_moved"], result["identity_drift_merged"]) == (0, 1)
    assert result["identity_drift_lost_decisions"] == 1
    # 경위 이름은 개정 2 에서 `merge_overwritten` → `row_replaced` 로 바뀌었다(ADR 0009 §5-2 (마)):
    # 이 조건이 잡는 **주** 경로에는 병합이 없어(실측 R1 `merged=0`) "merge" 라는 이름 자체가 거짓이었다.
    # `new_doc_id` 가 `None` 인 것은 "모른다"가 아니라 **다시 판단할 곳이 없다**는 사실이다.
    assert result["identity_drift"]["lost_decisions"] == [
        {"activity_id": ACTIVITY_MERGE, "doc_id": survivor, "decision": "confirmed", "cause": "row_replaced",
         "new_doc_id": None, "changed_fields": ["sender", "doc_number"], "approval_flipped": True}]

    reviews = _drift_reviews(client, auth, project_id)
    assert len(reviews) == 1, reviews
    assert reviews[0]["review_request_id"] == result["identity_drift_review_id"]
    assert reviews[0]["status"] == "open" and reviews[0]["assignee_role"] == "cm"


def test_v5_merge_overwritten_survives_as_a_live_row_with_someone_elses_content(client, auth, user_ids,
                                                                               merge_overwritten) -> None:
    """이 경로가 고아 판정에 걸리지 않는 이유를 그대로 고정한다: 행도 `reviewed_by` 도 살아 있고
    고아 표시조차 없다. 바뀐 것은 그 `doc_id` 가 담고 있는 **대장 행**이다."""
    project_id, survivor = merge_overwritten["project_id"], merge_overwritten["survivor"]
    detail = _document_detail(client, auth, project_id, survivor)
    assert detail["document"]["is_orphaned"] is False
    assert merge_overwritten["before_document"]["approval_status"] == "REJECTED"
    assert detail["document"]["approval_status"] == "APPROVED"                 # 다른 대장 행의 값이다
    assert detail["document"]["doc_number"] != merge_overwritten["before_document"]["doc_number"]

    mappings = [m for m in detail["mappings"] if m["activity_id"] == ACTIVITY_MERGE]
    assert len(mappings) == 1 and mappings[0]["needs_review"] is False
    assert mappings[0]["reviewed_by"] == user_ids["cm"]                        # 확정은 그대로 서 있다


def test_v5_merge_overwritten_flips_drawing_approval_but_no_longer_silently(client, auth, merge_overwritten) -> None:
    """**뒤집힘 자체는 의도된 동작이다** — 살아남은 행은 대장의 다른 행이고 대장이 정본이다(ADR 0007
    §1 규칙 1). 이번 사이클이 고친 것은 그것이 조용하다는 점이다. 그래서 두 사실을 **함께** 단언한다:
    `drawing_approval` 0.0 → 1.0(미승인 도면을 차단하던 근거가 사라졌다)이고, 같은 적재가 그 사건을
    CM 큐에 올렸다. 뒤집힘만 고정하면 큐가 사라져도 초록이고, 큐만 고정하면 이 결함이 무엇이었는지가
    테스트에서 사라진다."""
    before = merge_overwritten["before_readiness"]
    assert before["components"]["drawing_approval"] == 0.0
    assert [b["kind"] for b in _drawing_blockers(before)] == ["document_unapproved"]

    after = _readiness(client, auth, merge_overwritten["project_id"], ACTIVITY_MERGE)
    assert after["components"]["drawing_approval"] == 1.0
    assert not _drawing_blockers(after)                       # 착수 가능 쪽으로 열렸다
    assert after["score"] > before["score"]
    assert merge_overwritten["job"]["result"]["identity_drift_review_id"] is not None


# ── 양성 2: 확정이 **삼켜지는 행**에 있다 ────────────────────────────────────────
@pytest.fixture(scope="module")
def merge_absorbed(client, auth, user_ids, tmp_path_factory, merge_register) -> dict[str, Any]:
    """같은 병합의 반대편. CM 이 확정한 것이 **패자 쪽 행**(동부이앤씨)이면 그 `doc_id` 는 새 값을 얻지
    못한 채 사라진다 — 새 문서가 생기지 않으므로 고아 ↔ 신규 짝짓기로도 잡히지 않는다."""
    project_id = _new_project(client, auth, user_ids, "V5 병합 — 삼켜지는 행")
    upload(client, auth("contractor"), project_id, SCHEDULE)
    upload(client, auth("cm"), project_id, merge_register)

    doc_ids = _duct_doc_ids(client, auth, project_id)
    absorbed = doc_ids["동부이앤씨"]
    _resolve_mapping_review_for_doc(client, auth, project_id, ACTIVITY_MERGE, absorbed,
                                    "approved", "이 도면을 이 작업의 근거로 삼는다")
    before_readiness = _readiness(client, auth, project_id, ACTIVITY_MERGE)

    config_dir = _write_mutated_config(tmp_path_factory.mktemp("v5-absorbed"), _merge_sender_aliases)
    _, job = _upload_with_config(client, auth("cm"), project_id, merge_register, config_dir)
    return {"project_id": project_id, "job": job, "absorbed": absorbed, "survivor": doc_ids["동부건설"],
            "before_readiness": before_readiness}


def test_v5_merge_absorbed_puts_the_lost_decision_on_the_cm_queue(client, auth, merge_absorbed) -> None:
    project_id, absorbed = merge_absorbed["project_id"], merge_absorbed["absorbed"]
    result = merge_absorbed["job"]["result"]
    assert (result["identity_drift_moved"], result["identity_drift_merged"]) == (0, 1)
    assert result["identity_drift"]["lost_decisions"] == [
        {"activity_id": ACTIVITY_MERGE, "doc_id": absorbed, "decision": "confirmed", "cause": "row_absorbed",
         "new_doc_id": merge_absorbed["survivor"], "changed_fields": [], "approval_flipped": False}]
    assert result["orphaned_doc_ids"] == [absorbed]           # 흡수된 행은 이번 적재에 나타나지 않는다

    reviews = _drift_reviews(client, auth, project_id)
    assert len(reviews) == 1, reviews
    assert reviews[0]["review_request_id"] == result["identity_drift_review_id"]
    assert reviews[0]["status"] == "open" and reviews[0]["assignee_role"] == "cm"


def test_v5_merge_absorbed_loses_the_approval_evidence_conservatively(client, auth, merge_absorbed) -> None:
    """삼켜진 쪽은 §3 (가)와 같은 **보수적** 실패다 — 확정 매핑이 고아 문서를 가리켜 근거에서 빠지고
    점수가 내려간다. 위험한 것은 양성 1 쪽이지만, 이쪽도 사람의 판단이 오염된 것은 같으므로 큐에 오른다."""
    project_id = merge_absorbed["project_id"]
    assert _document_detail(client, auth, project_id, merge_absorbed["absorbed"])["document"]["is_orphaned"] is True
    assert _document_detail(client, auth, project_id, merge_absorbed["survivor"])["document"]["is_orphaned"] is False

    before, after = merge_absorbed["before_readiness"], _readiness(client, auth, project_id, ACTIVITY_MERGE)
    assert before["components"]["drawing_approval"] == 1.0
    assert after["components"]["drawing_approval"] < 1.0
    assert [b["kind"] for b in _drawing_blockers(after)] == ["document_mapping_pending"]


# ── 검토요청 문구: 약속의 **내용**을 단언한다(문자열 통째로가 아니라) ───────────────
# 이 저장소는 웹 테스트가 **거짓 문구를 계약으로 고정**해 169건이 전부 통과한 적이 있다. 그래서
# 여기서는 문장을 베껴 쓰지 않고 "그 경위에서 참일 수 없는 말이 없다"를 건다.
def test_v5_row_replaced_title_does_not_claim_orphaning_merging_or_movement(client, auth, merge_overwritten) -> None:
    """`row_replaced` 는 고아가 아니고(행이 살아 있다), 이동도 아니며(`moved == 0`), 다시 확정할
    **새 doc_id 자체가 없다**. 초판 문구는 이 셋을 모두 반대로 적었다.

    **"병합"도 쓰지 않는다(개정 2).** 이 적재는 실제로 `merged == 1` 이지만, 같은 경위가 발화하는 **주**
    경로(사명 변경 주 — 아래 V7a)는 `merged == 0` 이다. 여기서 "병합"을 계약으로 고정하면 병합이 없는
    적재에서 CM 이 있지도 않은 충돌 묶음을 찾게 되는 문구가 통과해 버린다(ADR 0009 §5-3 개정 2 정정 ②).
    """
    title = _drift_reviews(client, auth, merge_overwritten["project_id"])[0]["title"]
    assert "고아" not in title, title
    assert "병합" not in title, title
    assert "이동" not in title, title                 # "0건 이동했고" 를 포함해, 이동을 말하지 않는다
    assert "다시 확정" not in title, title             # 없는 행동을 시키지 않는다
    assert "1건" in title, title                      # 오염된 판단 건수는 사실대로 적는다
    assert "복구되지 않습니다" in title, title           # 매핑 복구를 약속하지 않는다(§5-3)


def test_v5_row_absorbed_title_does_not_claim_orphaning_or_merging(client, auth, merge_absorbed) -> None:
    title = _drift_reviews(client, auth, merge_absorbed["project_id"])[0]["title"]
    assert "고아" not in title, title
    assert "병합" not in title, title
    assert "이동" not in title, title
    assert "다시 확정" not in title, title
    assert "1건" in title, title
    assert "복구되지 않습니다" in title, title


def test_v5_row_moved_title_names_the_move_without_calling_it_orphaning(client, auth, sender_alias_drift) -> None:
    """대조군 — 경위가 `row_moved` 인 적재의 문구는 반대로 **이동과 다시 확정할 곳을 말해야 한다**.
    이것이 없으면 "세 경위 모두 어휘 몇 개를 빼면 통과"라는 틀린 구현이 위 두 테스트를 그대로 지나간다.

    **이 테스트는 개정 1 에서 "'고아'가 제목에 있어야 한다"고 단언했고, 그것이 거짓 계약이었다.**
    ADR 0009 §5-3 개정 2 정정 ①이 그 문장을 지목한다: 판정은 §5-2 (가)에서 이미 고아를 보지 않기로
    고쳤는데(좌변은 "이번 적재에 나타나지 않은 기존 행 전부") 문구만 고아라고 말하고 있었다. 실측이
    그것을 증명한다 — 워크북 시트명 변경 경로는 `moved=9` 인데 옛 행이 `is_orphaned=False` 다(P3).
    계획 0003 §12-d 도 `row_moved` 문구에 **"고아"라고 쓰지 않는다**고 명시한다. 이 저장소에는 웹 테스트
    169건이 거짓 문구를 계약으로 고정한 채 전원 통과한 전례가 있고(CLAUDE.md §6-4), 그 형태를 여기서
    되풀이하지 않는다. 그래서 이 테스트는 **판정이 실제로 관측한 값**(옮겨간 새 doc_id 가 있다 / 그
    위에서 다시 판단하라)만 요구하고, 판정이 보지 않는 사실(고아 여부·병합 여부)은 요구하지 않는다.
    """
    title = _drift_reviews(client, auth, sender_alias_drift["project_id"])[0]["title"]
    # 판정이 보지 않는 사실을 단정하지 않는다.
    assert "고아" not in title, title
    assert "병합" not in title, title
    # 관측한 값은 그대로 말한다 — 이 경위에만 있는 것: 이동 쌍, 옮겨간 새 doc_id, 그 위에서의 재확정.
    assert "이동" in title, title
    assert "새 doc_id" in title, title
    assert "다시 확정" in title, title
    assert "1건" in title, title
    assert "복구되지 않습니다" in title, title


# ── 음성 대조군 ───────────────────────────────────────────────────────────────
def test_v5_n1_register_result_update_without_collision_opens_no_review(client, auth, user_ids, tmp_path) -> None:
    """N1 — 충돌 **없이** 대장이 처리결과를 반려 → 승인으로 정상 갱신했다. `drawing_approval` 은 똑같이
    0.0 → 1.0 으로 뒤집히지만 이것은 사건이 아니다: **대장이 정본**이다(ADR 0007 §1 규칙 1).

    이 대조군이 없으면 "확정된 문서의 승인 상태가 바뀌면 발화"라는 틀린 구현이 통과하고, 그러면 대장이
    다음 주에 처리결과를 갱신할 때마다 CM 큐가 오염된다 — 그 끝은 운영자가 **탐지를 끄는 것**이다.
    """
    project_id = _new_project(client, auth, user_ids, "V5 N1 대장 정상 갱신")
    upload(client, auth("contractor"), project_id, SCHEDULE)
    upload(client, auth("cm"), project_id,
           _register_with_extra_tfa_rows(tmp_path / "before.xlsx", [_MERGE_ROWS[0]]))
    doc_id = _duct_doc_ids(client, auth, project_id)["동부건설"]
    _resolve_mapping_review_for_doc(client, auth, project_id, ACTIVITY_MERGE, doc_id, "approved", "반려 확인")
    assert _readiness(client, auth, project_id, ACTIVITY_MERGE)["components"]["drawing_approval"] == 0.0

    _, job = upload(client, auth("cm"), project_id, _register_with_extra_tfa_rows(
        tmp_path / "after.xlsx", [{**_MERGE_ROWS[0], "result": "승인"}]))

    assert (job["result"]["created"], job["result"]["orphaned"]) == (0, 0)
    assert job["result"]["identity_drift"] is None
    assert job["result"]["identity_drift_review_id"] is None
    assert not _has_warning(job, "DOCUMENT_IDENTITY_COLLISION")
    assert _drift_reviews(client, auth, project_id) == []
    # 뒤집힘은 일어난다 — 그리고 그것이 옳다. 사건은 "우리 규칙이 두 행을 뭉갰을 때"뿐이다.
    assert _readiness(client, auth, project_id, ACTIVITY_MERGE)["components"]["drawing_approval"] == 1.0


def test_v5_n1b_result_update_outside_the_collision_group_is_not_reported(client, auth, user_ids, tmp_path) -> None:
    """N1 강화판 — 같은 적재에 **관계없는 병합이 하나 있는** 상태에서 대장이 다른 문서의 처리결과를
    갱신한다. 드리프트 보고서 자체는 생기지만(merged 1건) 오염된 판단은 **0건**이어야 한다.

    N1 만으로는 오탐 방지 조건 ①("이번 적재의 충돌 묶음에 있을 것")을 고정하지 못한다: 그 조건을
    지워도 N1 에는 병합이 없어 `identity_drift` 가 통째로 `None` 이라 검토요청이 만들어지지 않는다.
    조건이 실제로 걸리는 자리는 "병합은 있는데 갱신된 문서는 그 묶음 밖"인 이 적재다.
    """
    project_id = _new_project(client, auth, user_ids, "V5 N1b 묶음 밖 갱신")
    upload(client, auth("contractor"), project_id, SCHEDULE)
    before = _register_with_extra_tfa_rows(tmp_path / "n1b_before.xlsx",
                                           [_MERGE_ROWS[0], *_STANDING_COLLISION_ROWS])
    _, first = upload(client, auth("cm"), project_id, before)
    assert first["result"]["identity_drift_merged"] == 1       # 상시 충돌 1건(사람 판단은 걸려 있지 않다)
    doc_id = _duct_doc_ids(client, auth, project_id)["동부건설"]
    _resolve_mapping_review_for_doc(client, auth, project_id, ACTIVITY_MERGE, doc_id, "approved", "반려 확인")

    _, job = upload(client, auth("cm"), project_id, _register_with_extra_tfa_rows(
        tmp_path / "n1b_after.xlsx", [{**_MERGE_ROWS[0], "result": "승인"}, *_STANDING_COLLISION_ROWS]))

    assert job["result"]["identity_drift_merged"] == 1         # 병합은 여전히 보고된다
    assert job["result"]["identity_drift"]["lost_decisions"] == []
    assert job["result"]["identity_drift_lost_decisions"] == 0
    assert job["result"]["identity_drift_review_id"] is None
    assert _drift_reviews(client, auth, project_id) == []


def test_v5_n2_reupload_after_the_merge_opens_no_new_review(client, auth, merge_overwritten) -> None:
    """N2 — 병합 뒤 **같은 config·같은 파일**을 다시 올린다. 충돌은 매 적재 다시 보고되지만 사건은 이미
    일어났고 승자도 그대로다 — 새 검토요청은 없다(사건이 일어난 적재에서 한 번만). 이것이 없으면
    현장의 주간 대장 업로드가 매주 같은 요청을 CM 큐에 쌓는 구현이 통과한다.

    첫 요청이 실제로 만들어졌다는 사실은 위 양성 1 테스트가 고정한다 — 여기서 그것을 다시 단언하면
    이 음성 테스트가 양성 1의 방어와 함께 무너져(뮤테이션이 두 곳을 동시에 빨갛게 만들어) "무엇이
    깨졌는가"를 가리키지 못한다. 이 테스트가 고정하는 것은 **적재 사이의 증분이 0**이라는 것뿐이다."""
    project_id = merge_overwritten["project_id"]
    before = _drift_reviews(client, auth, project_id)

    _, job = _upload_with_config(client, auth("cm"), project_id, merge_overwritten["register"],
                                 merge_overwritten["config_dir"])

    assert job["result"]["identity_drift_merged"] == 1         # 충돌 자체는 계속 보인다(대장이 그대로다)
    assert job["result"]["identity_drift"]["lost_decisions"] == []
    assert job["result"]["identity_drift_review_id"] is None
    assert _drift_reviews(client, auth, project_id) == before  # 큐가 자라지 않는다


def test_v5_n3_merge_without_human_decisions_warns_but_opens_no_review(
    client, auth, user_ids, tmp_path, merge_register,
) -> None:
    """N3 — 같은 별칭표 통합이지만 **사람의 판단이 하나도 없는** 프로젝트. 경고와 병합 보고는 뜨되
    검토요청은 만들지 않는다(ADR 0009 §5-2 큐 오염 방지 — V3b 가 첫 적재 충돌로 세운 계약을 config
    변경으로 생긴 병합에서도 그대로 유지한다)."""
    project_id = _new_project(client, auth, user_ids, "V5 N3 판단 없는 병합")
    upload(client, auth("contractor"), project_id, SCHEDULE)
    upload(client, auth("cm"), project_id, merge_register)
    assert not _reviews(client, auth, project_id, kind="document_mapping", status="approved")

    config_dir = _write_mutated_config(tmp_path / "cfg", _merge_sender_aliases)
    _, job = _upload_with_config(client, auth("cm"), project_id, merge_register, config_dir)

    assert _has_warning(job, "DOCUMENT_IDENTITY_COLLISION"), _warning_messages(job)
    assert job["result"]["identity_drift_merged"] == 1
    assert job["result"]["identity_drift"]["lost_decisions"] == []
    assert job["result"]["identity_drift_review_id"] is None
    assert _drift_reviews(client, auth, project_id) == []


# ═══════════════════════════════════════════════════════════════════════════
# V7 — 판정을 **행-정체 / 행-내용** 으로 다시 세운다 (ADR 0009 개정 2 / 계획 0003 §12-e)
#
# 개정 1 은 (나)·(다)를 "**한 적재 안에서** 두 개 이상의 대장 행이 같은 doc_id 로 수렴"(= 충돌 묶음)
# 위에 세웠다. 그 한정어 하나가 **사명 변경 주의 정상 운영**을 표 밖으로 밀어냈다: 별칭표를 통합하면서
# 대장에서 옛 법인명 행이 빠지면 두 행이 한 적재에 함께 있지 않아 묶음이 만들어지지 않는다. 실측 결과는
# 이 사이클이 이미 두 번 고친 blocker 와 데이터 모양·결과가 같았다 — `drawing_approval` 0.0 → 1.0
# (미승인 도면 위에서 착수 가능)인데 `identity_drift=None`, 검토요청 0건, DOCUMENT_IDENTITY_* 경고 0건.
#
# 아래 각 테스트의 docstring 첫 줄에 **"개정 1 코드에서 나오던 값"** 을 적어 둔다(계획 0003 §12-e).
# 그 값이 지금 값과 다르다는 것이 "이 시나리오가 결함을 잡는다"는 증거다(CLAUDE.md §6-2 규칙 1).
#
# 반증(계획 §12-e) — 이 단언들만으로는 잡지 못하는 것:
#  · `identity_drift_review_id is not None` **하나만** 걸면 무엇이든 발화하는 오탐 코드도 통과한다.
#    그래서 음성 V7e·V7f 가 같은 파일에 있다.
#  · `cause` 문자열만 걸면 **문구가 거짓인 채로** 통과한다. 그래서 문구 테스트는 문장을 베끼지 않고
#    "그 상황에서 참일 수 없는 말이 없다"를 건다(§6-4 규칙 3).
#  · V7a 에서 `drawing_approval` 만 걸면 **탐지가 사라져도 초록**이다(뒤집힘 자체는 의도된 동작이다).
# ═══════════════════════════════════════════════════════════════════════════
#: TFR 시트 열 번호(헤더 3행: No·문서발생일·발신·공종·번호·문서번호·제목·처리결과·처리완료일 — TFA 와 달리
#: `회신요청일` 이 없다). TFA 와 마찬가지로 **픽스처를 만드는 쪽**에서만 열 위치를 상수로 쓴다.
_TFR = {"no": 1, "issued": 2, "sender": 3, "discipline": 4, "seq": 5, "doc_number": 6,
        "title": 7, "result": 8, "completed": 9}
_TFR_FIRST_DATA_ROW = 4      # 헤더 3행 다음
_TFR_FIXTURE_ROW_COUNT = 2   # 픽스처가 담고 있는 TFR 데이터 행 수

#: 사명 변경 주(ADR 0009 §5-2 (바) R1). 같은 `번호`·같은 제목 원문에 **발신 표기만 다른** 두 행 —
#: 별칭표가 둘을 갈라 두는 동안에는 서로 다른 문서다. 1주차 대장에 둘 다 있고, 2주차에는 새 법인명 행만
#: 남으며 config 에는 `sender_aliases.동부건설 += 한빛이앤씨` 한 줄이 들어간다.
_OLD_NAME_ROW = {"sender": "동부", "discipline": "기계", "seq": _MERGE_SEQ, "result": "반려",
                 "doc_number": "동부-HG-TFA-기계-26-080", "title": _MERGE_TITLE}
_NEW_NAME_ROW = {"sender": "한빛이앤씨", "discipline": "기계", "seq": _MERGE_SEQ, "result": "승인",
                 "doc_number": "한빛-HG-TFA-기계-26-081", "title": _MERGE_TITLE}


def _write_register_row(ws, cols: dict[str, int], row_index: int, no: int, row: dict[str, Any]) -> None:
    ws.cell(row=row_index, column=cols["no"], value=no)
    ws.cell(row=row_index, column=cols["issued"], value=row.get("issued", "26-09-20"))
    ws.cell(row=row_index, column=cols["sender"], value=row.get("sender", "동부"))
    ws.cell(row=row_index, column=cols["discipline"], value=row.get("discipline", "구조"))
    ws.cell(row=row_index, column=cols["seq"], value=row["seq"])
    ws.cell(row=row_index, column=cols["doc_number"], value=row.get("doc_number"))
    ws.cell(row=row_index, column=cols["title"], value=row["title"])
    ws.cell(row=row_index, column=cols["result"], value=row.get("result"))


def _register_with_rows(dest: Path, *, tfa_rows: Sequence[dict[str, Any]] = (),
                        tfr_rows: Sequence[dict[str, Any]] | None = None,
                        drop_doc_number: bool = False) -> Path:
    """`_register_with_extra_tfa_rows` 의 확장판 — TFR 시트에도 행을 놓고, **문서번호 열 자체를 지운
    대장**(현장에 실제로 있다, ADR 0009 §5-2 (바) P9)을 만들 수 있다. 원본 픽스처는 건드리지 않는다.

    `tfr_rows` 를 주면 픽스처의 기존 TFR 데이터 2행을 **치우고** 그 자리에 쓴다(그 시트를 시나리오
    전용으로 쓰기 위해서다). `drop_doc_number` 는 두 시트에서 `문서번호` 열을 통째로 삭제한다 —
    파서는 헤더 별칭으로 열을 찾으므로 나머지 열은 그대로 읽힌다(ADR 0007 §2-5).
    """
    wb = openpyxl.load_workbook(REGISTER, data_only=True)
    for offset, row in enumerate(tfa_rows):
        _write_register_row(wb["TFA"], _TFA, _TFA_FIRST_FREE_ROW + offset, 90 + offset, row)
    if tfr_rows is not None:
        wb["TFR"].delete_rows(_TFR_FIRST_DATA_ROW, _TFR_FIXTURE_ROW_COUNT)
        for offset, row in enumerate(tfr_rows):
            _write_register_row(wb["TFR"], _TFR, _TFR_FIRST_DATA_ROW + offset, 90 + offset, row)
    if drop_doc_number:
        for sheet_name, cols in (("TFA", _TFA), ("TFR", _TFR)):
            wb[sheet_name].delete_cols(cols["doc_number"], 1)
    wb.save(dest)
    return dest


def _absorb_new_company_name(cfg: dict[str, Any]) -> None:
    """사명 변경 주에 실제로 하는 config 변경 — 새 법인명을 기존 표준명 아래로 넣는 **한 줄**이다
    (ADR 0009 §5-1: 동결할 수 없는 식별 표면). 대장 파일은 이 변경과 무관하게 정상적으로 갱신된다."""
    cfg["normalization"]["sender_aliases"]["동부건설"].append("한빛이앤씨")


def _rename_week_registers(tmp_path: Path, *, drop_doc_number: bool = False) -> tuple[Path, Path]:
    """(1주차, 2주차) 대장. 2주차에서 **옛 법인명 행이 빠진다** — 대장 쪽의 정상적인 주간 갱신이다."""
    suffix = "_nodocno" if drop_doc_number else ""
    return (_register_with_rows(tmp_path / f"week1{suffix}.xlsx", tfa_rows=[_OLD_NAME_ROW, _NEW_NAME_ROW],
                                drop_doc_number=drop_doc_number),
            _register_with_rows(tmp_path / f"week2{suffix}.xlsx", tfa_rows=[_NEW_NAME_ROW],
                                drop_doc_number=drop_doc_number))


def _company_rename_week(client, auth, user_ids, tmp_path: Path, name: str, *,
                         confirm_on: str, drop_doc_number: bool = False,
                         also_reject_on: str | None = None,
                         mutate_config=_absorb_new_company_name) -> dict[str, Any]:
    """사명 변경 주 재현. `confirm_on` 은 CM 이 검토 큐에서 확정할 행의 `sender_normalized` 다
    (`"동부건설"` = 살아남는 쪽 / `"한빛이앤씨"` = 사라지는 쪽). `mutate_config` 가 `None` 이면
    **config 를 바꾸지 않는다**(음성 대조군 V7b)."""
    week1, week2 = _rename_week_registers(tmp_path, drop_doc_number=drop_doc_number)
    project_id = _new_project(client, auth, user_ids, name)
    upload(client, auth("contractor"), project_id, SCHEDULE)
    _, first = upload(client, auth("cm"), project_id, week1)
    assert first["status"] == "done" and first["result"]["identity_drift"] is None   # 첫 적재는 판정하지 않는다

    doc_ids = _duct_doc_ids(client, auth, project_id)
    assert set(doc_ids) == {"동부건설", "한빛이앤씨"}, doc_ids   # 별칭표가 둘을 갈라 두고 있다
    _resolve_mapping_review_for_doc(client, auth, project_id, ACTIVITY_MERGE, doc_ids[confirm_on],
                                    "approved", "이 도면을 이 작업의 도면 근거로 삼는다")
    if also_reject_on is not None:
        _resolve_mapping_review_for_doc(client, auth, project_id, ACTIVITY_MERGE, doc_ids[also_reject_on],
                                        "rejected", "이 문서는 이 작업과 무관하다")
    before_document = _document_detail(client, auth, project_id, doc_ids[confirm_on])["document"]
    before_readiness = _readiness(client, auth, project_id, ACTIVITY_MERGE)

    if mutate_config is None:
        _, job = upload(client, auth("cm"), project_id, week2)
    else:
        config_dir = _write_mutated_config(tmp_path / f"cfg-{confirm_on}", mutate_config)
        _, job = _upload_with_config(client, auth("cm"), project_id, week2, config_dir)
    return {"project_id": project_id, "job": job, "survivor": doc_ids["동부건설"],
            "vanished": doc_ids["한빛이앤씨"], "before_document": before_document,
            "before_readiness": before_readiness}


# ── V7a / V7h — 판단이 **살아남는 쪽**에 있다(이번 blocker) ──────────────────────
@pytest.fixture(scope="module")
def company_rename_survivor(client, auth, user_ids, tmp_path_factory) -> dict[str, Any]:
    return _company_rename_week(client, auth, user_ids, tmp_path_factory.mktemp("v7a"),
                                "V7a 사명 변경 — 살아남는 쪽", confirm_on="동부건설")


def test_v7a_company_rename_week_reports_row_replaced_with_the_fields_that_changed(company_rename_survivor) -> None:
    """**개정 1 코드에서: `identity_drift=None`, 검토요청 없음, 경고 0건.** 충돌 묶음이 만들어지지 않아
    (나)·(다)의 전제가 거짓이었기 때문이다 — 두 행이 한 적재에 함께 있지 않다.

    지금은 `row_replaced` 1건으로 발화한다. `changed_fields` 가 **무엇이 달라졌는지**를 값으로 실어
    CM 이 "다른 문서로 바뀐 것"과 "대장이 오타를 고친 것"을 한 줄 안에서 가른다(ADR 0009 §5-2 (마)).
    """
    result = company_rename_survivor["job"]["result"]
    assert result["identity_drift"] is not None
    assert result["identity_drift"]["lost_decisions"] == [
        {"activity_id": ACTIVITY_MERGE, "doc_id": company_rename_survivor["survivor"],
         "decision": "confirmed", "cause": "row_replaced", "new_doc_id": None,
         "changed_fields": ["sender", "doc_number"], "approval_flipped": True}]
    assert result["identity_drift_lost_decisions"] == 1


def test_v7a_company_rename_week_flips_drawing_approval_and_opens_the_cm_review_together(
    client, auth, company_rename_survivor,
) -> None:
    """**개정 1 코드에서: `drawing_approval` 0.0 → 1.0 은 똑같이 일어나고 검토요청은 0건.** 그것이
    "미승인 도면 위에서 착수 가능"이 조용히 열린 모습이다(ADR 0009 §3 이 스스로 최악이라 적은 경로).

    §6-2 규칙 4 — **두 사실을 함께 단언한다.** 뒤집힘만 고정하면 큐가 사라져도 초록이고(뒤집힘 자체는
    대장이 정본이므로 의도된 동작이다), 큐만 고정하면 이 결함이 무엇이었는지가 테스트에서 사라진다.
    """
    project_id = company_rename_survivor["project_id"]
    before = company_rename_survivor["before_readiness"]
    assert before["components"]["drawing_approval"] == 0.0                    # 반려 도면이 차단하고 있었다
    assert [b["kind"] for b in _drawing_blockers(before)] == ["document_unapproved"]
    assert company_rename_survivor["before_document"]["approval_status"] == "REJECTED"

    after = _readiness(client, auth, project_id, ACTIVITY_MERGE)
    assert after["components"]["drawing_approval"] == 1.0                     # 착수 가능 쪽으로 열렸다
    assert not _drawing_blockers(after)
    # …그리고 같은 적재가 그 사건을 CM 큐에 올렸다.
    reviews = _drift_reviews(client, auth, project_id)
    assert len(reviews) == 1, reviews
    assert reviews[0]["review_request_id"] == company_rename_survivor["job"]["result"]["identity_drift_review_id"]
    assert reviews[0]["status"] == "open" and reviews[0]["assignee_role"] == "cm"


def test_v7h_gate_regression_no_move_and_no_merge_but_the_review_still_opens(company_rename_survivor) -> None:
    """**게이트 회귀.** `IdentityDriftReport` 를 만드는 게이트가 `moved or merged` 로 되돌아가면 판정이
    옳게 발화해도 보고서가 만들어지지 않아 검토요청이 **다시 조용히 삼켜진다**(실측: 새 조건 + 옛 게이트
    = `identity_drift=None`, 요청 0건 — 고치기 전과 완전히 같다).

    그래서 "`moved`·`merged` 가 **둘 다 0**"과 "그런데도 검토요청이 있다"를 **한 단언 안에서** 고정한다.
    이 두 줄이 없으면 게이트를 되돌려도 이 파일은 초록이다(계획 0003 §12-e V7h).
    """
    result = company_rename_survivor["job"]["result"]
    assert (result["identity_drift_moved"], result["identity_drift_merged"]) == (0, 0)
    assert result["identity_drift_review_id"] is not None


def test_v7a_title_reports_the_flip_and_the_changed_fields_without_claiming_a_merge(
    client, auth, company_rename_survivor,
) -> None:
    """문구 — `row_replaced` 의 **주** 경로다(`merged == 0`). 여기서 "병합"이라고 적으면 CM 은 있지도
    않은 충돌 묶음을 찾는다(ADR 0009 §5-3 개정 2 정정 ②). "고아"·"이동"도 이 적재에서는 참이 아니고,
    다시 확정할 새 `doc_id` 도 없다. 대신 **관측한 값**은 반드시 말해야 한다: 승인 근거가 뒤집혔다는
    사실과, 달라진 행-정체 필드가 무엇인지(`changed_fields`).
    """
    title = _drift_reviews(client, auth, company_rename_survivor["project_id"])[0]["title"]
    for forbidden in ("고아", "병합", "이동", "다시 확정"):
        assert forbidden not in title, title
    assert "뒤집혔습니다" in title, title            # approval_flipped=True 를 문구가 실제로 쓴다
    assert "발신" in title and "문서번호" in title, title   # changed_fields 를 값으로 나열한다
    assert "복구되지 않습니다" in title, title


def test_v7a_title_tail_points_at_the_config_because_the_fingerprint_moved(
    client, auth, company_rename_survivor,
) -> None:
    """문구의 **꼬리는 지문에서 유도된다**(ADR 0009 §5-2 서두: 지문은 판정 조건이 아니라 "어디를
    되돌려야 하는가"를 답하는 보고 값이다). 이 경로는 config 를 실제로 바꿨으므로 config 를 되돌리라고
    적어야 하고, 지문이 그것을 증명한다. 아래 시트명 경로 테스트가 **반대쪽**을 고정한다 — 둘이 함께
    있어야 "언제나 config 라고 쓰는" 구현이 걸린다.
    """
    drift = company_rename_survivor["job"]["result"]["identity_drift"]
    assert drift["previous_fingerprint"] != drift["current_fingerprint"]
    title = _drift_reviews(client, auth, company_rename_survivor["project_id"])[0]["title"]
    assert "config 가 바뀌었습니다" in title, title
    assert "되돌리" in title, title


# ── V7b — 같은 삭제, **config 만 안 바꿈**(음성 대조군) ──────────────────────────
def test_v7b_same_deletion_without_the_config_change_is_a_conservative_orphan(
    client, auth, user_ids, tmp_path,
) -> None:
    """**개정 1 코드에서도 같았다**(ADR 0009 §5-2 (바) R2) — 이 대조군은 "삭제가 있으면 발화"라는 틀린
    구현을 막는다. 대장에서 옛 법인명 행이 빠지는 것은 같지만 우리 식별 규칙은 움직이지 않았으므로,
    확정은 **고아 문서를 가리키는 보수적 실패**로 끝난다(0.0 → 0.5, 관측 가능). 사건이 아니다.
    """
    fixture = _company_rename_week(client, auth, user_ids, tmp_path, "V7b 대조군 — config 무변",
                                   confirm_on="동부건설", mutate_config=None)
    project_id, result = fixture["project_id"], fixture["job"]["result"]
    assert result["identity_drift"] is None
    assert result["identity_drift_review_id"] is None
    assert _drift_reviews(client, auth, project_id) == []
    assert not _has_warning(fixture["job"], "DOCUMENT_IDENTITY_DRIFT")
    assert result["orphaned_doc_ids"] == [fixture["survivor"]]
    assert _document_detail(client, auth, project_id, fixture["survivor"])["document"]["is_orphaned"] is True

    after = _readiness(client, auth, project_id, ACTIVITY_MERGE)
    assert fixture["before_readiness"]["components"]["drawing_approval"] == 0.0
    assert after["components"]["drawing_approval"] == 0.5      # 보수적으로 내려간다(뒤집히지 않는다)


# ── V7c — 판단이 **사라지는 쪽**에 있다(대칭 짝) ────────────────────────────────
@pytest.fixture(scope="module")
def company_rename_vanished(client, auth, user_ids, tmp_path_factory) -> dict[str, Any]:
    return _company_rename_week(client, auth, user_ids, tmp_path_factory.mktemp("v7c"),
                                "V7c 사명 변경 — 사라지는 쪽", confirm_on="한빛이앤씨")


def test_v7c_symmetric_pair_reports_row_absorbed_and_names_where_the_row_went(
    client, auth, company_rename_vanished,
) -> None:
    """**개정 1 코드에서: 침묵**(고아만 되고 검토요청 0건). 이쪽은 §3 (가)와 같은 보수적 실패라 점수는
    내려가지만, 사람의 판단이 오염된 것은 살아남는 쪽과 똑같다 — 그래서 큐에 오른다.

    `new_doc_id` 가 **살아남은 doc_id 를 가리킨다**는 것이 이 경위의 핵심이다(`row_replaced` 와 달리
    다시 판단할 곳이 실제로 있다). 그 값을 드리프트 보고서가 아니라 **1주차에 관측한 doc_id** 와
    비교한다 — 보고서끼리 비교하면 "보고서가 자기 자신과 같다"만 남는다.
    """
    project_id, result = company_rename_vanished["project_id"], company_rename_vanished["job"]["result"]
    assert result["identity_drift"]["lost_decisions"] == [
        {"activity_id": ACTIVITY_MERGE, "doc_id": company_rename_vanished["vanished"],
         "decision": "confirmed", "cause": "row_absorbed",
         "new_doc_id": company_rename_vanished["survivor"], "changed_fields": [], "approval_flipped": False}]
    assert (result["identity_drift_moved"], result["identity_drift_merged"]) == (0, 0)
    assert result["orphaned_doc_ids"] == [company_rename_vanished["vanished"]]

    reviews = _drift_reviews(client, auth, project_id)
    assert len(reviews) == 1 and reviews[0]["review_request_id"] == result["identity_drift_review_id"]
    before = company_rename_vanished["before_readiness"]["components"]["drawing_approval"]
    assert before == 1.0
    assert _readiness(client, auth, project_id, ACTIVITY_MERGE)["components"]["drawing_approval"] < before


def test_v7c_title_points_at_the_doc_id_that_now_holds_the_row(client, auth, company_rename_vanished) -> None:
    """문구 — 이 경위도 `merged == 0` 이므로 "병합"이라고 쓸 수 없고, 이동 쌍도 없다. 대신 **다시 판단할
    곳이 있다**는 사실은 반드시 적어야 한다(`new_doc_id` 가 차 있다)."""
    title = _drift_reviews(client, auth, company_rename_vanished["project_id"])[0]["title"]
    for forbidden in ("고아", "병합", "이동"):
        assert forbidden not in title, title
    assert "다른 문서" in title and "다시 판단" in title, title
    assert "복구되지 않습니다" in title, title


# ── V7d — 같은 사건을 **문서번호 열이 없는 대장**에서 ────────────────────────────
def test_v7d_company_rename_week_fires_even_without_a_doc_number_column(client, auth, user_ids, tmp_path) -> None:
    """**개정 1 코드에서: 침묵**(ADR 0009 §5-2 (바) P9). 문서번호 열이 없는 현장이 실제로 있다.

    행-정체가 `(sender, doc_number, seq_raw, title)` 에서 사실상 세 필드로 줄어도 **발신 원문이 두 행을
    갈라 준다** — 그래서 `changed_fields` 가 `["sender"]` 하나로 줄 뿐 발화는 유지된다. 이 테스트가
    없으면 "문서번호가 함께 바뀔 때만 발화"하는 구현이 통과한다.
    """
    fixture = _company_rename_week(client, auth, user_ids, tmp_path, "V7d 문서번호 열 없는 대장",
                                   confirm_on="동부건설", drop_doc_number=True)
    result = fixture["job"]["result"]
    assert all(d["doc_number"] is None for d in _documents(client, auth, fixture["project_id"])), "열이 지워졌다"
    assert result["identity_drift"]["lost_decisions"] == [
        {"activity_id": ACTIVITY_MERGE, "doc_id": fixture["survivor"], "decision": "confirmed",
         "cause": "row_replaced", "new_doc_id": None, "changed_fields": ["sender"], "approval_flipped": True}]
    assert (result["identity_drift_moved"], result["identity_drift_merged"]) == (0, 0)
    assert result["identity_drift_review_id"] is not None
    assert fixture["before_readiness"]["components"]["drawing_approval"] == 0.0
    assert _readiness(client, auth, fixture["project_id"], ACTIVITY_MERGE)["components"]["drawing_approval"] == 1.0


# ── V7g — 행-정체까지 같은 두 행 + `sheet_doc_types` 병합 ((나-ii) 전용) ──────────
def test_v7g_identical_rows_merged_by_sheet_doc_types_still_fire_through_row_content(
    client, auth, user_ids, tmp_path,
) -> None:
    """**개정 1 코드에서: 발화(1건).** 즉 이 시나리오는 개정 2 가 **잃지 말아야 할** 경로다 —
    ADR 0009 §5-2 (바) P13, 역방향 확인(CLAUDE.md §6-3)이 잡아낸 구멍이다.

    문서번호 열이 없는 대장에서 **행-정체 네 필드가 모두 같은** 두 행이 시트 둘에 나뉘어 있으면
    `sheet_doc_types` 를 한 줄 고쳐 둘을 한 `doc_id` 로 합쳐도 (나-i)는 침묵한다(행-정체가 달라지지
    않았으므로). 잡아내는 것은 **(나-ii)** 뿐이다: 이 `doc_id` 가 다른 `doc_id` 를 흡수했고 행-내용
    (처리결과·승인 상태)이 달라졌다. `changed_fields == []` 가 그 사실을 값으로 말한다.

    **(나-ii)를 빼면 실패해야 하는 시나리오다.** 조건이 합집합이어야 하는 이유가 여기 있다.
    """
    project_id = _new_project(client, auth, user_ids, "V7g 행-정체 동일 + 시트 병합")
    upload(client, auth("contractor"), project_id, SCHEDULE)
    same_row = {"sender": "동부", "discipline": "기계", "seq": _MERGE_SEQ, "title": _MERGE_TITLE}
    register = _register_with_rows(tmp_path / "identical_rows.xlsx",
                                   tfa_rows=[{**same_row, "result": "반려"}],
                                   tfr_rows=[{**same_row, "result": "승인"}], drop_doc_number=True)
    _, first = upload(client, auth("cm"), project_id, register)
    assert first["result"]["identity_drift"] is None
    duct = {d["doc_type"]: d for d in _documents(client, auth, project_id) if d["title"] == _MERGE_TITLE}
    assert set(duct) == {"TFA", "TFR"}, duct              # 시트가 갈라 두는 동안에는 서로 다른 문서다
    tfa_doc_id = duct["TFA"]["doc_id"]
    _resolve_mapping_review_for_doc(client, auth, project_id, ACTIVITY_MERGE, tfa_doc_id,
                                    "approved", "반려된 도면임을 확인 — 이 작업의 도면 근거로 삼는다")
    before = _readiness(client, auth, project_id, ACTIVITY_MERGE)
    assert before["components"]["drawing_approval"] == 0.0

    config_dir = _write_mutated_config(
        tmp_path / "cfg", lambda cfg: cfg["register_layout"]["sheet_doc_types"]["TFA"].append("TFR"))
    _, job = _upload_with_config(client, auth("cm"), project_id, register, config_dir)

    result = job["result"]
    assert result["identity_drift"]["lost_decisions"] == [
        {"activity_id": ACTIVITY_MERGE, "doc_id": tfa_doc_id, "decision": "confirmed",
         "cause": "row_replaced", "new_doc_id": None, "changed_fields": [], "approval_flipped": True}]
    assert result["identity_drift_review_id"] is not None
    assert result["identity_drift_merged"] == 1           # 이쪽은 실제로 한 적재 안의 충돌이다
    assert _readiness(client, auth, project_id, ACTIVITY_MERGE)["components"]["drawing_approval"] == 1.0


def test_v7g_title_says_the_raw_register_fields_are_unchanged(client, auth, user_ids, tmp_path) -> None:
    """문구 — `changed_fields == []` 인 경위다. 여기서 "발신이 달라졌습니다" 류를 적으면 **관측하지 못한
    것을 단정**하는 것이다(대장 원문 네 필드는 그대로다). 이 적재는 `merged == 1` 이지만 그렇다고
    "병합"을 문구의 계약으로 고정하지는 않는다 — 같은 경위의 주 경로(V7a)는 `merged == 0` 이고, 그때
    같은 문장이 거짓이 된다(ADR 0009 §5-3).
    """
    project_id = _new_project(client, auth, user_ids, "V7g 문구")
    upload(client, auth("contractor"), project_id, SCHEDULE)
    same_row = {"sender": "동부", "discipline": "기계", "seq": _MERGE_SEQ, "title": _MERGE_TITLE}
    register = _register_with_rows(tmp_path / "identical_rows.xlsx",
                                   tfa_rows=[{**same_row, "result": "반려"}],
                                   tfr_rows=[{**same_row, "result": "승인"}], drop_doc_number=True)
    upload(client, auth("cm"), project_id, register)
    tfa_doc_id = [d["doc_id"] for d in _documents(client, auth, project_id)
                  if d["title"] == _MERGE_TITLE and d["doc_type"] == "TFA"][0]
    _resolve_mapping_review_for_doc(client, auth, project_id, ACTIVITY_MERGE, tfa_doc_id, "approved", "반려 확인")
    config_dir = _write_mutated_config(
        tmp_path / "cfg", lambda cfg: cfg["register_layout"]["sheet_doc_types"]["TFA"].append("TFR"))
    _upload_with_config(client, auth("cm"), project_id, register, config_dir)

    title = _drift_reviews(client, auth, project_id)[0]["title"]
    for forbidden in ("고아", "병합", "이동", "다시 확정"):
        assert forbidden not in title, title
    assert "그대로인데" in title, title              # 대장 원문 네 필드가 안 바뀌었다는 사실을 적는다
    assert "뒤집혔습니다" in title, title            # approval_flipped=True


# ── V7i — 블라인드 스팟 실측(§6-1): `column_aliases.sender` 로 **열 자체**를 옮긴다 ──
@pytest.fixture(scope="module")
def sender_column_alias_drift(client, auth, user_ids, tmp_path_factory) -> dict[str, Any]:
    """ADR 0009 §5-2 (바) "이 기준이 놓치는 것" 3 — **적어 두기만 하고 태우지 않은 스팟**이다.
    계획 0003 §12-e V7i 가 qa 에게 "결과를 먼저 관측하고 그 값을 단언으로 적으라"고 지시한다."""
    project_id = _prepared_project(client, auth, user_ids, "V7i 발신 열 별칭 변경")
    review = _resolve_mapping_review(client, auth, project_id, ACTIVITY_CONFIRM, "approved", "확정")
    confirmed_doc_id = review["conflicting_sources"]["doc_id"]
    confirmed_before = _document_detail(client, auth, project_id, confirmed_doc_id)["document"]
    config_dir = _write_mutated_config(
        tmp_path_factory.mktemp("v7i"),
        # 발신을 **다른 열에서 읽게** 만든다. 대장 파일은 한 바이트도 바뀌지 않았다.
        lambda cfg: cfg["register_layout"]["column_aliases"].__setitem__("sender", ["공종"]))
    _, job = _upload_with_config(client, auth("cm"), project_id, REGISTER, config_dir)
    moved_to = [d["doc_id"] for d in _documents(client, auth, project_id)
                if d["doc_number"] == confirmed_before["doc_number"] and d["title"] == confirmed_before["title"]
                and d["doc_id"] != confirmed_doc_id]
    assert len(moved_to) == 1, moved_to
    return {"project_id": project_id, "job": job, "confirmed_doc_id": confirmed_doc_id,
            "moved_to_doc_id": moved_to[0]}


def test_v7i_column_alias_change_moves_every_row_and_is_reported_as_row_moved(sender_column_alias_drift) -> None:
    """**개정 1 코드에서: 미실측**(ADR 0009 §5-2 (바)가 "실측하지 않았다"고 적어 둔 세 스팟 중 하나).
    실제로 태운 결과를 그대로 단언으로 고정한다 — 문서 10건 전부가 새 `doc_id` 를 얻고((가)가 제목
    원문으로 짝지어 `moved=10`), 옛 행 10건이 고아가 되며, 사람의 판단 1건이 `row_moved` 로 걸린다.

    (다)는 이 경로를 잡지 못한다(대장 원문에서 읽는 **열 자체**가 바뀌어 행-정체가 통째로 달라진다).
    그것이 ADR 이 남긴 구멍이고, 여기서는 **(가)가 덮는다**는 사실을 값으로 남긴다.
    """
    result = sender_column_alias_drift["job"]["result"]
    assert (result["created"], result["orphaned"]) == (FIXTURE_DOCUMENT_COUNT, FIXTURE_DOCUMENT_COUNT)
    assert result["identity_drift_moved"] == FIXTURE_DOCUMENT_COUNT
    assert result["identity_drift_merged"] == 0
    assert result["identity_drift"]["lost_decisions"] == [
        {"activity_id": ACTIVITY_CONFIRM, "doc_id": sender_column_alias_drift["confirmed_doc_id"],
         "decision": "confirmed", "cause": "row_moved",
         "new_doc_id": sender_column_alias_drift["moved_to_doc_id"],
         "changed_fields": [], "approval_flipped": False}]
    assert result["identity_drift_review_id"] is not None


# ── V7e·V7f — 개정 1 의 **오탐 둘**이 음성이 됐다 ────────────────────────────────
#: 첫 적재부터 같은 `doc_id` 로 수렴하는 두 행(= 충돌이 **상시화된** 대장). 매주 같은 두 행이 올라온다.
_STANDING_PAIR = [
    {"sender": "동부", "discipline": "기계", "seq": 26090, "doc_number": "동부-HG-TFA-기계-26-090",
     "title": _MERGE_TITLE, "result": "승인"},
    {"sender": "동부", "discipline": "기계", "seq": 26090, "doc_number": "동부-HG-TFA-기계-26-091",
     "title": _MERGE_TITLE, "result": "반려"},
]


def test_v7e_result_update_inside_a_standing_collision_group_is_not_reported(
    client, auth, user_ids, tmp_path,
) -> None:
    """**개정 1 코드에서: 오탐 1건 + 거짓 문구**(ADR 0009 §5-2 (바) P4 / MINOR-1).

    충돌이 상시화된 대장에서는 개정 1 의 조건 ①("이번 적재의 충돌 묶음에 있을 것")이 **항상 참**이라
    가드가 행 지문 하나만 남았고, 대장이 그 행의 처리결과를 정상 갱신할 때마다 "내용이 **다른 대장 행**
    으로 바뀌었습니다"라는 거짓 문구와 함께 CM 큐가 오염됐다. 실제로는 **같은 행**의 처리결과 갱신이다.

    `test_v5_n1b_*` 와 다르다: 저쪽은 갱신된 문서가 묶음 **밖**이라 개정 1 의 조건 ①이 거짓이었다.
    조건 ①이 실제로 걸리던 자리는 갱신이 묶음 **안**에서 일어나는 이 적재다.
    """
    project_id = _new_project(client, auth, user_ids, "V7e 상시 충돌 안의 정상 갱신")
    upload(client, auth("contractor"), project_id, SCHEDULE)
    _, first = upload(client, auth("cm"), project_id,
                      _register_with_rows(tmp_path / "v7e_before.xlsx", tfa_rows=_STANDING_PAIR))
    assert first["result"]["identity_drift_merged"] == 1        # 첫 적재부터 상시 충돌이다
    doc_id = _duct_doc_ids(client, auth, project_id)["동부건설"]
    _resolve_mapping_review_for_doc(client, auth, project_id, ACTIVITY_MERGE, doc_id, "approved", "반려 확인")
    assert _readiness(client, auth, project_id, ACTIVITY_MERGE)["components"]["drawing_approval"] == 0.0

    # 대장이 **같은 행**(승자)의 처리결과를 반려 → 승인으로 정상 갱신했다. 우리 규칙은 그대로다.
    _, job = upload(client, auth("cm"), project_id, _register_with_rows(
        tmp_path / "v7e_after.xlsx",
        tfa_rows=[_STANDING_PAIR[0], {**_STANDING_PAIR[1], "result": "승인"}]))

    assert job["result"]["identity_drift_merged"] == 1          # COLLISION 은 그대로 보고된다
    assert _has_warning(job, "DOCUMENT_IDENTITY_COLLISION"), _warning_messages(job)
    assert job["result"]["identity_drift"]["lost_decisions"] == []
    assert job["result"]["identity_drift_lost_decisions"] == 0
    assert job["result"]["identity_drift_review_id"] is None
    assert _drift_reviews(client, auth, project_id) == []
    # 뒤집힘은 일어난다 — 대장이 정본이다(ADR 0007 §1 규칙 1). 사건이 아니다.
    assert _readiness(client, auth, project_id, ACTIVITY_MERGE)["components"]["drawing_approval"] == 1.0


def test_v7f_real_deletion_of_a_blank_doc_number_row_is_not_reported_as_absorbed(
    client, auth, user_ids, tmp_path,
) -> None:
    """**개정 1 코드에서: 오탐 1건**(ADR 0009 §5-2 (바) P5 / MINOR-2).

    개정 1 의 흡수 짝짓기는 "충돌 묶음 구성원과 **제목이 같고 문서번호가 호환**된다"였고, 그 호환 판별은
    한쪽이 비면 통과시킨다(문서번호 열이 없는 현장을 위한 완화). 그래서 문서번호가 **빈 행**은 "제목만
    같으면 통과"로 퇴화해, 무관한 충돌이 하나 있는 적재에서 **진짜 삭제**가 흡수로 오보고됐다.

    행-정체 **전체 일치**를 요구하면 이 오탐이 사라진다 — 진짜로 지워진 행의 행-정체는 이번 적재
    어디에도 없기 때문이다. 남는 것은 고아 표시뿐이고, 그것이 옳다.
    """
    project_id = _new_project(client, auth, user_ids, "V7f 문서번호 빈 행의 진짜 삭제")
    upload(client, auth("contractor"), project_id, SCHEDULE)
    #: 충돌 묶음 구성원과 **제목만 같고** 발신·번호가 다르며 문서번호가 비어 있는 행.
    blank_doc_number_row = {"sender": "중원", "discipline": "기계", "seq": 26095, "doc_number": None,
                            "title": _MERGE_TITLE, "result": "승인"}
    upload(client, auth("cm"), project_id, _register_with_rows(
        tmp_path / "v7f_before.xlsx", tfa_rows=[*_STANDING_PAIR, blank_doc_number_row]))
    doc_id = _duct_doc_ids(client, auth, project_id)["중원엔지니어링"]
    _resolve_mapping_review_for_doc(client, auth, project_id, ACTIVITY_MERGE, doc_id,
                                    "approved", "이 도면을 이 작업의 근거로 삼는다")

    # 그 행을 대장에서 **진짜로 지운다**. 무관한 충돌 묶음은 그대로 남아 있다.
    _, job = upload(client, auth("cm"), project_id,
                    _register_with_rows(tmp_path / "v7f_after.xlsx", tfa_rows=_STANDING_PAIR))

    assert job["result"]["identity_drift_merged"] == 1          # 무관한 충돌은 계속 보고된다
    assert job["result"]["orphaned_doc_ids"] == [doc_id]        # 남는 것은 고아 표시뿐이다
    assert job["result"]["identity_drift"]["lost_decisions"] == []
    assert job["result"]["identity_drift_review_id"] is None
    assert _drift_reviews(client, auth, project_id) == []


# ── 경위가 **섞인** 적재의 문구 ─────────────────────────────────────────────────
def test_v7_mixed_causes_title_writes_each_cause_side_by_side(client, auth, user_ids, tmp_path) -> None:
    """한 적재가 두 경위를 함께 만든다(살아남는 쪽 확정 + 사라지는 쪽 반려). 경위를 **합치면 거짓이
    된다**(ADR 0009 §5-3) — `row_replaced` 에는 다시 판단할 곳이 없고 `row_absorbed` 에는 있다.

    그래서 두 절이 **나란히** 적히는지를 고정한다: 뒤집힘·달라진 필드(앞 절)와 "지금은 다른 문서 아래에
    있으니 그 위에서 다시 판단하라"(뒤 절)가 한 제목에 함께 있어야 하고, 반려된 판단을 "다시 확정"
    하라고 시키지 않아야 한다(`_redecide_verb` 가 값에서 유도한다).
    """
    fixture = _company_rename_week(client, auth, user_ids, tmp_path, "V7 경위 혼합",
                                   confirm_on="동부건설", also_reject_on="한빛이앤씨")
    lost = fixture["job"]["result"]["identity_drift"]["lost_decisions"]
    assert [(x["cause"], x["decision"]) for x in lost] == [("row_replaced", "confirmed"),
                                                           ("row_absorbed", "rejected")]
    title = _drift_reviews(client, auth, fixture["project_id"])[0]["title"]
    for forbidden in ("고아", "병합", "이동"):
        assert forbidden not in title, title
    assert "뒤집혔습니다" in title and "발신" in title, title          # row_replaced 절
    assert "다른 문서" in title and "다시 판단" in title, title        # row_absorbed 절
    assert "또한" in title, title                                    # 두 절이 합쳐지지 않고 나란히 적힌다
    assert "다시 확정" not in title, title            # 반려된 판단을 확정하라고 시키지 않는다


# ── 문구의 꼬리는 **지문에서 유도된다** — config 를 안 바꾼 경로 ───────────────────
def test_v7_sheet_rename_title_tail_does_not_tell_the_cm_to_revert_a_config(
    client, auth, user_ids, tmp_path,
) -> None:
    """워크북 시트명 변경은 config 를 **한 글자도 바꾸지 않는다**(`fingerprint_changed=False`, 실측).
    그런데도 꼬리에 "config 를 되돌리십시오"라고 적으면 CM 은 바뀐 적 없는 config 를 뒤지게 되고,
    진짜 원인(대장 파일 쪽 입력)은 문구 밖에 남는다. **무조건 config 라고 쓰는 구현**이 이 테스트에서만
    걸리도록, 위 V7a 의 반대쪽(지문이 바뀐 경로)과 짝으로 둔다.
    """
    project_id = _prepared_project(client, auth, user_ids, "V7 시트명 문구")
    _resolve_mapping_review(client, auth, project_id, ACTIVITY_CONFIRM, "approved", "확정")
    wb = openpyxl.load_workbook(REGISTER, data_only=True)
    wb["TFA"].title = "자료제출"
    renamed = tmp_path / "sheet_renamed_title.xlsx"
    wb.save(renamed)
    _, job = upload(client, auth("cm"), project_id, renamed)

    drift = job["result"]["identity_drift"]
    assert drift["previous_fingerprint"] == drift["current_fingerprint"], "config 는 바뀌지 않았다"
    title = _drift_reviews(client, auth, project_id)[0]["title"]
    assert "config 는 그대로입니다" in title, title
    assert "config 가 바뀌었습니다" not in title, title
    assert "되돌리" not in title, title              # 되돌릴 config 가 없다
    assert "대장 파일" in title, title               # 어디를 봐야 하는지는 말한다
