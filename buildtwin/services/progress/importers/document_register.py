"""문서관리대장(xlsx) → `Document` 목록 파서 (ADR 0007 §2). 순수 함수 — DB 쓰기는 이 모듈의 범위 밖이다.

컬럼 위치·시트명·처리결과 표기·공종/발신 별칭은 전부 `config/document_register.yaml` 에서 읽는다
(CLAUDE.md §3 규칙 5·10). 코드에는 열 위치("3행", "H열") 도 상수로 두지 않는다 — 헤더 행을
`register_layout.header_row_search_range` 안에서 `column_aliases` 로 탐색해 매 시트마다 새로 찾는다
(ADR 0007 §2-5).

안전 규칙(ADR 0007 §3-2): 처리결과가 공란이면 `UNKNOWN`/confidence 1.0/`register_status_blank`.
비어 있지 않은데 어떤 규칙에도 안 걸리면 `UNKNOWN`/confidence 0.0/`register_status_unmatched`/
`needs_review=True`. 두 경우 모두 "승인"으로 추측하지 않는다 — 값이 다른 것은 운영 신호일 뿐
안전성(둘 다 UNKNOWN)에는 영향이 없다.

`doc_number`(대장에서 수식으로 파생되는 컬럼)는 절대 되파싱하지 않는다(§2-4). 구조화된 값
(`sender`/`discipline_raw`/`seq_raw`)은 언제나 대장의 해당 컬럼에서 읽고, `doc_number`가 그 값들과
어긋나 보이면 경고만 남기고 컬럼 값을 신뢰한다.
"""
from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from packages.core.models.document import Document, DocumentApprovalStatus, DocumentType
from packages.core.models.evidence import Evidence

from ..config_loader import load_config

_CONFIG_FILENAME = "document_register.yaml"


# ─────────────────────────────────────────────────────────────────────────────
# 반환 타입 — Document 매핑·Activity 매핑은 이 모듈의 범위 밖이므로 여기서는 파싱 결과만 담는다.
# ─────────────────────────────────────────────────────────────────────────────
@dataclass
class RegisterWarning:
    """`config/document_register.yaml` `import_warnings` 의 code 중 하나. 메시지는 영어로 고정하고
    (CLAUDE.md §3 규칙 5: 식별자는 영어), 시트명·문서번호 등 대장에서 온 값만 그대로 끼워 넣는다 —
    그 값이 한국어여도 코드가 만든 리터럴이 아니라 데이터이므로 규칙 5 위반이 아니다."""

    code: str
    detail: str
    sheet: str | None = None
    row: int | None = None

    def __str__(self) -> str:
        loc = f" [{self.sheet}#{self.row}]" if self.sheet and self.row else (f" [{self.sheet}]" if self.sheet else "")
        return f"{self.code}{loc}: {self.detail}"


@dataclass
class DocumentRegisterImportResult:
    """대장 xlsx 한 파일을 파싱한 결과. DB 조회가 필요한 §2-2 규칙 2·4(orphan·renamed 판정)는
    포함하지 않는다 — 이 함수는 순수 파싱까지다. `is_orphaned` 는 항상 False 로 반환한다."""

    documents: list[Document] = field(default_factory=list)
    warnings: list[RegisterWarning] = field(default_factory=list)
    sheet_counts: dict[str, int] = field(default_factory=dict)   # sheet_name -> 적재한 문서 수

    @property
    def warning_messages(self) -> list[str]:
        """`JobRow.warnings`(list[str])에 그대로 넣을 수 있는 문자열 뷰."""
        return [str(w) for w in self.warnings]


@dataclass
class _StatusRule:
    id: str
    status: str
    confidence: float
    pattern: re.Pattern[str]


# ─────────────────────────────────────────────────────────────────────────────
# config 로딩
# ─────────────────────────────────────────────────────────────────────────────
def _load_register_config() -> dict[str, Any]:
    return load_config(_CONFIG_FILENAME)


def _compile_status_rules(rules_cfg: list[dict[str, Any]]) -> list[_StatusRule]:
    return [
        _StatusRule(
            id=str(r["id"]), status=str(r["status"]), confidence=float(r["confidence"]),
            pattern=re.compile(str(r["pattern"]), re.IGNORECASE),
        )
        for r in rules_cfg
    ]


# ─────────────────────────────────────────────────────────────────────────────
# 헤더 행 탐색 (ADR 0007 §2-5) — 열 위치를 상수로 두지 않는다
# ─────────────────────────────────────────────────────────────────────────────
def _normalize_header_text(value: Any) -> str:
    return re.sub(r"\s+", "", str(value)).strip().lower()


def _match_columns_for_row(row_cells: list[tuple[str, int]], alias_norms: dict[str, list[str]]) -> dict[str, int]:
    """한 헤더 후보 행에 대해 논리 컬럼 → 열 번호 매핑을 만든다.

    "첫 일치를 쓴다"(config 주석)는 열 순서가 아니라 **별칭 우선순위**로 읽는다: 논리 컬럼마다
    자신의 별칭 목록을 순서대로 훑어 정확히 일치하는 첫 열을 취한다. 열 순서로 먼저 훑으면, 예를 들어
    행번호용 "No" 열이 `seq_raw` 별칭 목록의 낮은 우선순위 별칭("no")과 우연히 일치해 진짜 "번호" 열보다
    먼저 채간다 — 대장마다 다른 위치의 컬럼을 별칭만으로 찾는(§2-5) 설계에서는 별칭 우선순위가 이겨야
    한다. 이미 다른 논리 컬럼에 배정된 열은 재사용하지 않는다.
    """
    col_map: dict[str, int] = {}
    used_columns: set[int] = set()
    for logical, norms in alias_norms.items():
        for alias_norm in norms:
            hit = next((col for norm_text, col in row_cells if col not in used_columns and norm_text == alias_norm), None)
            if hit is not None:
                col_map[logical] = hit
                used_columns.add(hit)
                break
    return col_map


def _find_header(ws: Worksheet, layout: dict[str, Any]) -> tuple[int | None, dict[str, int]]:
    """`header_row_search_range` 안에서 `column_aliases` 가 가장 많이 일치하는 행을 헤더로 본다."""
    search_start, search_end = layout["header_row_search_range"]
    min_matched = int(layout["header_min_matched_columns"])
    alias_norms: dict[str, list[str]] = {
        logical: [_normalize_header_text(a) for a in aliases]
        for logical, aliases in layout["column_aliases"].items()
    }
    last_row = min(int(search_end), ws.max_row or 0)
    best_row: int | None = None
    best_map: dict[str, int] = {}
    for row_idx in range(int(search_start), last_row + 1):
        row_cells = [(_normalize_header_text(cell.value), cell.column) for cell in ws[row_idx] if cell.value is not None]
        col_map = _match_columns_for_row(row_cells, alias_norms)
        if len(col_map) > len(best_map):
            best_map, best_row = col_map, row_idx
    if best_row is None or len(best_map) < min_matched:
        return None, {}
    return best_row, best_map


def _sheet_doc_type(sheet_name: str, sheet_doc_types: dict[str, list[str]], fallback: str) -> str:
    lname = sheet_name.lower()
    for doc_type, aliases in sheet_doc_types.items():
        if any(alias.lower() in lname for alias in aliases):
            return doc_type
    return fallback


def _is_skip_sheet(sheet_name: str, skip_sheets: list[str]) -> bool:
    lname = sheet_name.lower()
    return any(alias.lower() in lname for alias in skip_sheets)


# ─────────────────────────────────────────────────────────────────────────────
# 값 정규화 (ADR 0007 §2-3)
# ─────────────────────────────────────────────────────────────────────────────
def _strip_or_none(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _seq_normalized(seq_raw: str | None) -> str | None:
    """숫자 이외 문자를 모두 제거해 이어붙인다. 자릿수를 재해석하지 않는다(연도 확장·0 제거 금지)."""
    if not seq_raw:
        return None
    digits = "".join(ch for ch in seq_raw if ch.isdigit())
    return digits or None


def _squash(text: str) -> str:
    """공백 제거 + 대문자화. `sender_normalized` 재료이므로 안정적이어야 한다."""
    return re.sub(r"\s+", "", text).upper()


def _build_sender_alias_lookup(sender_aliases: dict[str, list[str]]) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for canonical, aliases in sender_aliases.items():
        lookup[_squash(canonical)] = canonical
        for alias in aliases:
            lookup[_squash(alias)] = canonical
    return lookup


def _sender_normalized(sender_raw: str | None, alias_lookup: dict[str, str]) -> str:
    squashed = _squash(sender_raw or "")
    return alias_lookup.get(squashed, squashed)


def _discipline_normalized(discipline_raw: str | None, discipline_aliases: dict[str, list[str]]) -> str | None:
    """대장 공종은 신뢰 불가 필드다 — 못 찾으면 None(가점도 감점도 없음), 코드가 값을 추측하지 않는다."""
    if not discipline_raw:
        return None
    text = discipline_raw.strip()
    for canonical, aliases in discipline_aliases.items():
        if text in aliases:
            return canonical
    return None


def _title_normalized(title: str, normalize_cfg: dict[str, Any]) -> str:
    text = title
    for pattern in normalize_cfg.get("strip_patterns", []):
        text = re.sub(pattern, "", text, flags=re.IGNORECASE)
    strip_chars = normalize_cfg.get("strip_chars", "")
    if strip_chars:
        text = re.sub(f"[{re.escape(strip_chars)}]", " ", text)
    if normalize_cfg.get("lowercase", True):
        text = text.lower()
    if normalize_cfg.get("collapse_whitespace", True):
        text = re.sub(r"\s+", " ", text).strip()
    return text


_GENERIC_DATE_RE = re.compile(r"^(\d{2,4})[.\-/](\d{1,2})[.\-/](\d{1,2})$")


def _parse_cell_date(value: Any, date_formats: list[str]) -> str | None:
    """엑셀 날짜 셀은 그대로 date 로 읽고, 문자열 셀만 `date_formats` 로 시도한다(config 주석 그대로).

    셀 타입이 시트 안에서 혼재하는 경우(문자열/`datetime`)를 다뤄야 하므로 두 갈래를 모두 처리한다.
    `date_formats` 어디에도 안 맞는 문자열은(예: 2자리 연도+대시 구분자처럼 config 목록에 없는 변형)
    버리지 않고 일반적인 `y[-.]m[-.]d` 패턴으로 한 번 더 시도한 뒤, 그래도 안 되면 원문을 그대로
    보존한다 — 대장 값을 조용히 지우는 것보다 안전하다.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in date_formats:
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    m = _GENERIC_DATE_RE.match(text)
    if m:
        y, mo, d = m.groups()
        year = int(y) if len(y) == 4 else 2000 + int(y)
        try:
            return date(year, int(mo), int(d)).isoformat()
        except ValueError:
            return text
    return text


def _normalize_status(
    result_raw: str | None, rules: list[_StatusRule], blank_cfg: dict[str, Any], unmatched_cfg: dict[str, Any],
) -> tuple[str, float, bool, str, str | None]:
    """returns (status, confidence, needs_review, evidence.method, rule_id). ADR 0007 §3-2 표 그대로.

    판정 대상은 처리결과 칸을 strip 한 텍스트다(공백만 있는 칸도 공란으로 본다) — `result_raw`
    필드 자체는 별도로 원문 그대로 보관하므로 여기서 strip 해도 원문 손실이 아니다.
    """
    stripped = (result_raw or "").strip()
    if not stripped:
        return (str(blank_cfg["status"]), float(blank_cfg["confidence"]), bool(blank_cfg["needs_review"]),
                "register_status_blank", None)
    for rule in rules:
        if rule.pattern.search(stripped):
            return (rule.status, rule.confidence, False, "register_status_rule", rule.id)
    return (str(unmatched_cfg["status"]), float(unmatched_cfg["confidence"]), bool(unmatched_cfg["needs_review"]),
            "register_status_unmatched", None)


def _compute_doc_id(doc_type: str, sender_normalized: str, seq_normalized: str | None, title_normalized: str) -> str:
    """ADR 0007 §2-1. `discipline` 은 재료에 들어가지 않는다 — 신뢰 불가 필드가 문서 정체성에
    관여하면 협력사가 공종을 고쳐 적을 때 같은 문서가 다른 문서가 된다."""
    material = f"{doc_type}|{sender_normalized}|{seq_normalized or ''}|{title_normalized}"
    digest = hashlib.sha256(material.encode("utf-8")).hexdigest()[:16]
    return f"doc-{digest}"


def _doc_type_enum(doc_type: str) -> DocumentType:
    try:
        return DocumentType(doc_type)
    except ValueError:
        return DocumentType.OTHER


# ─────────────────────────────────────────────────────────────────────────────
# 행 → Document
# ─────────────────────────────────────────────────────────────────────────────
def _build_document(
    raw: dict[str, Any], title_val: str, sheet_name: str, doc_type: str, row_idx: int,
    project_id: str, file_id: str, file_uri: str | None,
    sender_alias_lookup: dict[str, str], discipline_aliases: dict[str, list[str]], date_formats: list[str],
    status_rules: list[_StatusRule], blank_cfg: dict[str, Any], unmatched_cfg: dict[str, Any],
    title_normalize_cfg: dict[str, Any], seen_doc_numbers: dict[str, list[tuple[str, int]]],
    warnings: list[RegisterWarning],
) -> Document:
    sender_raw = _strip_or_none(raw.get("sender")) or ""
    discipline_raw = _strip_or_none(raw.get("discipline_raw"))
    seq_raw = _strip_or_none(raw.get("seq_raw"))
    doc_number = _strip_or_none(raw.get("doc_number"))
    result_cell = raw.get("result_raw")
    # §3-2 규칙 3 / evidence.note: result_raw 는 원문 그대로 보관한다(앞뒤 공백 포함) — 정규화는
    # 별도 필드(approval_status)에만 반영하고 원문을 지우거나 덮어쓰지 않는다.
    result_raw = None if result_cell is None else str(result_cell)

    sender_normalized = _sender_normalized(sender_raw, sender_alias_lookup)
    discipline_normalized = _discipline_normalized(discipline_raw, discipline_aliases)
    seq_normalized = _seq_normalized(seq_raw)
    title_normalized = _title_normalized(title_val, title_normalize_cfg)

    issued_on = _parse_cell_date(raw.get("issued_on"), date_formats)
    completed_on = _parse_cell_date(raw.get("completed_on"), date_formats)

    status, confidence, needs_review, method, rule_id = _normalize_status(result_raw, status_rules, blank_cfg, unmatched_cfg)
    if method == "register_status_unmatched":
        warnings.append(RegisterWarning(
            "document_status_unmatched", f"doc_number={doc_number!r} title={title_val!r}",
            sheet=sheet_name, row=row_idx,
        ))

    if doc_number:
        seen_doc_numbers.setdefault(doc_number, []).append((sheet_name, row_idx))
        # ADR 0007 §2-4: doc_number 를 되파싱하지 않는다. 여기서는 구조를 복원하지 않고, 컬럼 원문이
        # doc_number 문자열 안에 부분 문자열로 나타나는지만 대조해 어긋남을 "경고"로만 남긴다.
        mismatched = [
            name for name, value in (("sender", sender_raw), ("discipline", discipline_raw))
            if value and value not in doc_number
        ]
        if mismatched:
            warnings.append(RegisterWarning(
                "doc_number_mismatch", f"doc_number={doc_number!r} mismatched_columns={mismatched}",
                sheet=sheet_name, row=row_idx,
            ))

    evidence = Evidence(
        source_type="document",
        source_id=file_id,
        file_uri=file_uri,
        method=method,
        rule_id=rule_id,
        note=result_raw,
        extra={"sheet": sheet_name, "row": row_idx, "doc_number": doc_number},
    )

    doc_id = _compute_doc_id(doc_type, sender_normalized, seq_normalized, title_normalized)

    return Document(
        project_id=project_id,
        doc_id=doc_id,
        doc_type=_doc_type_enum(doc_type),
        sender=sender_raw,
        sender_normalized=sender_normalized,
        discipline_raw=discipline_raw,
        discipline_normalized=discipline_normalized,
        seq_raw=seq_raw,
        seq_normalized=seq_normalized,
        doc_number=doc_number,
        title=title_val,
        title_normalized=title_normalized,
        issued_on=issued_on,
        result_raw=result_raw,
        approval_status=DocumentApprovalStatus(status),
        approval_confidence=confidence,
        approval_evidence=evidence,
        completed_on=completed_on,
        file_id=file_id,
        sheet_name=sheet_name,
        source_row=row_idx,
        needs_review=needs_review,
        is_orphaned=False,   # §2-2 규칙 2: 대장 전체(DB 기존 행)와 비교해야 하므로 이 순수 파서 밖의 일
    )


def _import_sheet(
    ws: Worksheet, sheet_name: str, doc_type: str, header_row: int, col_map: dict[str, int], layout: dict[str, Any],
    project_id: str, file_id: str, file_uri: str | None,
    sender_alias_lookup: dict[str, str], discipline_aliases: dict[str, list[str]], date_formats: list[str],
    status_rules: list[_StatusRule], blank_cfg: dict[str, Any], unmatched_cfg: dict[str, Any],
    title_normalize_cfg: dict[str, Any], seen_doc_numbers: dict[str, list[tuple[str, int]]],
    result: DocumentRegisterImportResult,
) -> int:
    data_start = header_row + int(layout.get("data_start_offset", 1))
    stop_streak = int(layout.get("blank_row_stop_streak", 20))
    last_row = ws.max_row or 0
    blank_streak = 0
    count = 0
    row_idx = data_start
    while row_idx <= last_row:
        raw = {logical: ws.cell(row=row_idx, column=col_idx).value for logical, col_idx in col_map.items()}
        title_val = _strip_or_none(raw.get("title"))
        if title_val is None:
            # 서식만 있고 값이 없는 행. `max_row` 를 그대로 데이터 끝으로 믿지 않고, 연속 공백 행이
            # `blank_row_stop_streak` 에 닿아야만 시트 스캔을 끝낸다(고립된 빈 행 한 줄로는 끝내지 않는다).
            blank_streak += 1
            if blank_streak >= stop_streak:
                break
            row_idx += 1
            continue
        blank_streak = 0
        doc = _build_document(
            raw, title_val, sheet_name, doc_type, row_idx, project_id, file_id, file_uri,
            sender_alias_lookup, discipline_aliases, date_formats, status_rules, blank_cfg, unmatched_cfg,
            title_normalize_cfg, seen_doc_numbers, result.warnings,
        )
        result.documents.append(doc)
        count += 1
        row_idx += 1
    return count


# ─────────────────────────────────────────────────────────────────────────────
# 진입점
# ─────────────────────────────────────────────────────────────────────────────
def import_document_register(
    path: str | Path, project_id: str, file_id: str, file_uri: str | None = None,
) -> DocumentRegisterImportResult:
    """대장 xlsx 를 읽어 `Document` 목록으로 정규화한다. DB 를 읽거나 쓰지 않는 순수 함수다.

    `is_orphaned` 판정(§2-2 규칙 2)과 `document_possibly_renamed` 경고(§2-2 규칙 4)는 기존 DB 상태와의
    비교가 필요하므로 이 함수의 범위 밖이다 — 호출자(persistence 계층)가 이 결과와 기존 `documents`
    테이블을 대조해 채운다.
    """
    cfg = _load_register_config()
    layout = cfg["register_layout"]
    normalization = cfg.get("normalization", {})
    status_rules = _compile_status_rules(cfg.get("status_normalization", []))
    blank_cfg = cfg["blank"]
    unmatched_cfg = cfg["unmatched"]
    title_normalize_cfg = cfg.get("title_matching", {}).get("normalize", {})

    sender_alias_lookup = _build_sender_alias_lookup(normalization.get("sender_aliases", {}))
    discipline_aliases = normalization.get("discipline_aliases", {})
    date_formats = list(normalization.get("date_formats", []))

    result = DocumentRegisterImportResult()
    seen_doc_numbers: dict[str, list[tuple[str, int]]] = {}

    wb = openpyxl.load_workbook(Path(path), data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if _is_skip_sheet(sheet_name, layout.get("skip_sheets", [])):
                result.warnings.append(RegisterWarning("sheet_skipped", "matched skip_sheets alias", sheet=sheet_name))
                continue

            header_row, col_map = _find_header(ws, layout)
            if header_row is None:
                result.warnings.append(RegisterWarning(
                    "header_row_not_found", f"no row in range matched >= {layout['header_min_matched_columns']} column_aliases",
                    sheet=sheet_name,
                ))
                continue

            missing_required = [c for c in layout.get("required_columns", ["title"]) if c not in col_map]
            if missing_required:
                result.warnings.append(RegisterWarning(
                    "required_column_missing", f"missing={missing_required}", sheet=sheet_name,
                ))
                continue

            doc_type = _sheet_doc_type(sheet_name, layout.get("sheet_doc_types", {}), str(layout.get("fallback_doc_type", "other")))
            count = _import_sheet(
                ws, sheet_name, doc_type, header_row, col_map, layout, project_id, file_id, file_uri,
                sender_alias_lookup, discipline_aliases, date_formats, status_rules, blank_cfg, unmatched_cfg,
                title_normalize_cfg, seen_doc_numbers, result,
            )
            result.sheet_counts[sheet_name] = count
    finally:
        wb.close()

    for doc_number, occurrences in seen_doc_numbers.items():
        if len(occurrences) > 1:
            locs = ", ".join(f"{s}#{r}" for s, r in occurrences)
            result.warnings.append(RegisterWarning("duplicate_doc_number", f"doc_number={doc_number!r} at {locs}"))

    return result


__all__ = ["DocumentRegisterImportResult", "RegisterWarning", "import_document_register"]
